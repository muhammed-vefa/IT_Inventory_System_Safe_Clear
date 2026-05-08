from flask import Blueprint, jsonify, request
from core.database_sql import query_db, get_db_connection
from core.auth import require_auth, require_editor, require_admin
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

depot_manager_bp = Blueprint('depot_manager', __name__)


@depot_manager_bp.route('/get_all', methods=['GET'])
@require_auth
def get_all():
    """Tüm depo envanterini getirir."""
    try:
        items = query_db("SELECT * FROM depot_items ORDER BY name ASC")
        return jsonify([dict(row) for row in items])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@depot_manager_bp.route('/weekly_report', methods=['GET'])
@require_auth
def get_weekly_report():
    """Haftalık depo hareketlerini ve güncel durumu getirir."""
    import datetime
    try:
        conn = get_db_connection()
        # Son 7 günün tarihi
        last_week = datetime.datetime.now() - datetime.timedelta(days=7)
        # SQL Server tarih formatı için string'e çevirelim
        last_week_str = last_week.strftime('%Y-%m-%d %H:%M:%S')

        items = conn.execute("SELECT * FROM depot_items ORDER BY category, name").fetchall()
        transactions = conn.execute("""
            SELECT t.*, i.name as item_name, i.category as item_category
            FROM depot_transactions t 
            JOIN depot_items i ON t.depot_item_id = i.id 
            WHERE t.created_at >= ? 
            ORDER BY t.created_at DESC
        """, (last_week_str,)).fetchall()
        
        return jsonify({
            "items": [dict(r) for r in items] if items else [],
            "transactions": [dict(r) for r in transactions] if transactions else []
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@depot_manager_bp.route('/export_weekly_excel', methods=['GET'])
@require_auth
def export_weekly_excel():
    """Haftalık depo raporunu Excel formatında üretir."""
    import datetime
    from flask import send_file
    import os

    try:
        conn = get_db_connection()
        # Verileri kategori ve isme göre sıralı al
        items = conn.execute("SELECT * FROM depot_items ORDER BY category, name").fetchall()
        
        # Son 7 günün hareketleri
        last_week = datetime.datetime.now() - datetime.timedelta(days=7)
        last_week_str = last_week.strftime('%Y-%m-%d %H:%M:%S')
        transactions = conn.execute("""
            SELECT t.*, i.name as item_name, i.category as item_category
            FROM depot_transactions t 
            JOIN depot_items i ON t.depot_item_id = i.id 
            WHERE t.created_at >= ? 
            ORDER BY i.category, i.name, t.created_at DESC
        """, (last_week_str,)).fetchall()
        conn.close()

        wb = openpyxl.Workbook()
        
        # --- 1. STOK DURUMU SAYFASI ---
        ws1 = wb.active
        ws1.title = "Güncel Stok Durumu"
        
        # Stil tanımları
        header_font = Font(bold=True, size=10, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell_font = Font(size=9)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')
        wrap_align = Alignment(vertical='center', wrap_text=True)

        headers = ["Kategori", "Ürün Adı", "Mevcut Stok", "Birim", "Kritik Sınır", "Açıklama"]
        for col, h in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col)
            cell.value = h
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        for row_idx, item in enumerate(items, 2):
            vals = [item['category'], item['name'], item['current_stock'], item['unit'], item['critical_stock'], item['description']]
            for col_idx, val in enumerate(vals, 1):
                cell = ws1.cell(row=row_idx, column=col_idx)
                cell.value = val
                cell.font = cell_font
                cell.border = border
                if col_idx == 2: # Ürün Adı
                    cell.alignment = wrap_align
                else:
                    cell.alignment = Alignment(vertical='center')

        # Sütun Genişlikleri
        ws1.column_dimensions['A'].width = 15 # Kategori
        ws1.column_dimensions['B'].width = 30 # Ürün Adı (Daraltıldı ve Wrap Text uygulandı)
        ws1.column_dimensions['C'].width = 12 # Stok
        ws1.column_dimensions['D'].width = 10 # Birim
        ws1.column_dimensions['E'].width = 12 # Kritik
        ws1.column_dimensions['F'].width = 40 # Açıklama

        # --- 2. HAREKETLER SAYFASI ---
        ws2 = wb.create_sheet("Haftalık Hareketler")
        headers2 = ["Tarih", "Kategori", "Ürün Adı", "İşlem", "Miktar", "Personel", "Not"]
        for col, h in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col)
            cell.value = h
            cell.font = header_font
            cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
            cell.alignment = center_align
            cell.border = border

        for row_idx, t in enumerate(transactions, 2):
            t_date = t['created_at']
            if isinstance(t_date, str): 
                try: t_date = datetime.datetime.strptime(t_date.split('.')[0], '%Y-%m-%d %H:%M:%S').strftime('%d.%m.%Y %H:%M')
                except: pass
            elif isinstance(t_date, datetime.datetime):
                t_date = t_date.strftime('%d.%m.%Y %H:%M')

            ttype = "Giriş (+)" if t['transaction_type'] == 'in' else "Çıkış (-)"
            vals = [t_date, t['item_category'], t['item_name'], ttype, t['quantity'], t['user_name'], t['note']]
            for col_idx, val in enumerate(vals, 1):
                cell = ws2.cell(row=row_idx, column=col_idx)
                cell.value = val
                cell.font = cell_font
                cell.border = border
                if col_idx == 3: # Ürün Adı
                    cell.alignment = wrap_align
                else:
                    cell.alignment = Alignment(vertical='center')

        ws2.column_dimensions['A'].width = 18
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 30
        ws2.column_dimensions['D'].width = 12
        ws2.column_dimensions['E'].width = 10
        ws2.column_dimensions['F'].width = 15
        ws2.column_dimensions['G'].width = 30

        filename = f"Depo_Haftalik_Rapor_{datetime.datetime.now().strftime('%d_%m_%Y')}.xlsx"
        save_path = os.path.join(os.getcwd(), filename)
        wb.save(save_path)
        
        return send_file(save_path, as_attachment=True)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@depot_manager_bp.route('/reset_weekly', methods=['POST'])
@require_editor
def reset_weekly_stats():
    """Haftalık dağıtım miktarını sıfırla ve Devreden değerini güncelle."""
    try:
        conn = get_db_connection()
        # Saha_stock'u "Devreden" (geçen haftadan kalan) olarak güncelle
        conn.execute("UPDATE depot_items SET saha_stock = current_stock, weekly_distributed = 0 WHERE category IN ('Sarf Malzeme', 'Gıda')")
        conn.commit()
        conn.close()
        return jsonify({"message": "Haftalık istatistikler sıfırlandı."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@depot_manager_bp.route('/alerts', methods=['GET'])
@require_auth
def get_alerts():
    """Kritik stok seviyesinin altındaki ürünleri getirir."""
    try:
        items = query_db("SELECT * FROM depot_items WHERE current_stock <= critical_stock")
        return jsonify([dict(row) for row in items])
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@depot_manager_bp.route('/add', methods=['POST'])
@require_editor
def add_item():
    """Depoya yeni ürün ekler."""
    data = request.json
    try:
        conn = get_db_connection()
        conn.execute(
            "INSERT INTO depot_items (category, name, critical_stock, current_stock, unit, description) VALUES (?,?,?,?,?,?)",
            (data.get('category'), data.get('name'), data.get('critical_stock', 5),
             data.get('current_stock', 0), data.get('unit', 'Adet'), data.get('description', ''))
        )
        conn.commit()
        conn.close()
        return jsonify({"message": "Ürün depoya eklendi"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@depot_manager_bp.route('/update/<int:item_id>', methods=['PUT'])
@require_editor
def update_item(item_id):
    """Depo ürününü günceller."""
    data = request.json
    try:
        conn = get_db_connection()
        conn.execute(
            "UPDATE depot_items SET category=?, name=?, critical_stock=?, current_stock=?, unit=?, description=?, saha_stock=?, arizali_stock=?, kayip_stock=? WHERE id=?",
            (data.get('category'), data.get('name'), data.get('critical_stock'),
             data.get('current_stock'), data.get('unit'), data.get('description'),
             data.get('saha_stock', 0), data.get('arizali_stock', 0), data.get('kayip_stock', 0), item_id)
        )
        conn.commit()
        conn.close()
        return jsonify({"message": "Ürün güncellendi"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def _norm(k):
    if not k: return ""
    s = str(k).upper().replace('İ','I').replace('Ğ','G').replace('Ü','U').replace('Ş','S').replace('Ö','O').replace('Ç','C')
    return " ".join(s.replace('\n', ' ').replace('\r', '').split()).strip()

def _get_val(item, item_keys, variants):
    for v in variants:
        v_n = _norm(v)
        if v_n in item_keys: return item[item_keys[v_n]]
    return None

def _clean_int(val, default=0):
    try:
        if val is None or str(val).strip() == "": return default
        return int(float(str(val).replace(',', '.')))
    except: return default

@depot_manager_bp.route('/sync_from_excel', methods=['POST'])
@require_editor
def sync_from_excel():
    import os
    from core.excel_utils import read_excel_data
    try:
        # Excel dosyasını bul
        excel_path = os.path.join('database', 'depo_envanter.xlsx')
        if not os.path.exists(excel_path):
            excel_path = os.path.join('database', 'depo_envanteri.xlsx')
        if not os.path.exists(excel_path):
            return jsonify({"error": "Excel dosyası bulunamadı (depo_envanter.xlsx veya depo_envanteri.xlsx)."}), 404

        conn = get_db_connection()
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        added_count = 0
        updated_count = 0
        dedup_map = {}  # (norm_name, category) -> data

        # Tüm sekmeleri tara (sekme adı = kategori ipucu)
        for s_idx, sheet in enumerate(wb.worksheets):
            try:
                data = read_excel_data(excel_path, sheet_name=s_idx)
            except Exception:
                continue
            if not data:
                continue

            print(f"DEBUG: Depo '{sheet.title}' sekmesi, {len(data)} satır.")

            for item in data:
                try:
                    item_keys = {_norm(k): k for k in item.keys()}

                    # Gerçek sütun adı: 'ÜRÜN ADI'
                    name = str(_get_val(item, item_keys, [
                        'URUN ADI', 'URUN', 'AD', 'NAME', 'MALZEME', 'MALZEME ADI'
                    ]) or '').strip()
                    if not name or name in ('0', '-', 'NONE', 'None'):
                        continue

                    # Gerçek sütun adı: 'KATEGORİ'
                    raw_cat = str(_get_val(item, item_keys, [
                        'KATEGORI', 'TURU', 'TUR', 'CATEGORY', 'GRUP', 'KATEGORİ', 'TÜR', 'TÜRÜ'
                    ]) or sheet.title or '').strip()

                    cat_norm = _norm(raw_cat)
                    if 'AG' in cat_norm or 'ALTYAPI' in cat_norm: category = 'AĞ VE ALTYAPI'
                    elif 'AKSESUAR' in cat_norm:                  category = 'AKSESUAR'
                    elif 'DONANIM' in cat_norm:                   category = 'DONANIM'
                    elif 'SARF' in cat_norm:                      category = 'SARF MALZEME'
                    elif 'GIDA' in cat_norm:                      category = 'GIDA'
                    elif 'KABLO' in cat_norm:                     category = 'KABLO'
                    else:
                        category = raw_cat if raw_cat else 'GENEL'

                    # Alan Eşleştirmeleri
                    current = _clean_int(_get_val(item, item_keys, [
                        'MEVCUT STOK', 'MEVCUT', 'KALAN', 'STOK', 'STOCK', 'DEPODA', 'DEPO', 'ADET', 'KALAN ADET'
                    ]))
                    toplam_val = _clean_int(_get_val(item, item_keys, ['TOPLAM', 'TOTAL']))
                    if current == 0 and toplam_val > 0:
                        current = toplam_val

                    critical = _clean_int(_get_val(item, item_keys, [
                        'KRITIK STOK', 'KRITIK', 'CRITICAL', 'MIN STOK', 'KRİTİK STOK', 'KRİTİK', 'KRITIK SEVIYE'
                    ]), 5)
                    saha    = _clean_int(_get_val(item, item_keys, ['SAHADA', 'SAHA', 'FIELD', 'SAHA STOK', 'ŞAHADA']))
                    arizali = _clean_int(_get_val(item, item_keys, ['ARIZALI', 'BOZUK', 'BROKEN', 'FAULTY', 'ARIZALI STOK']))
                    kayip   = _clean_int(_get_val(item, item_keys, ['KAYIP', 'LOST', 'MISSING', 'KAYIP STOK']))
                    unit    = str(_get_val(item, item_keys, ['BIRIM', 'UNIT', 'OLCU', 'BİRİM']) or 'Adet').strip()
                    desc    = str(_get_val(item, item_keys, ['ACIKLAMA', 'NOT', 'NOTLAR', 'DESCRIPTION', 'AÇIKLAMA', 'ACIKLAMALAR']) or '').strip()

                    key = (_norm(name), category)
                    if key in dedup_map:
                        dedup_map[key]['current'] += current
                    else:
                        dedup_map[key] = {
                            'category': category, 'name': name,
                            'current': current, 'critical': critical,
                            'unit': unit, 'description': desc,
                            'saha': saha, 'arizali': arizali, 'kayip': kayip
                        }
                except Exception as row_err:
                    print(f"Depot row error (sheet={sheet.title}): {row_err}")
                    continue

        # UPSERT: Mevcut varsa güncelle, yoksa ekle (manuel eklemeler korunur)
        for key, d in dedup_map.items():
            exists = conn.execute(
                "SELECT id FROM depot_items WHERE name=? AND category=?",
                (d['name'], d['category'])
            ).fetchone()
            if exists:
                conn.execute(
                    """UPDATE depot_items SET current_stock=?, critical_stock=?, saha_stock=?,
                       arizali_stock=?, kayip_stock=?, unit=?, description=? WHERE id=?""",
                    (d['current'], d['critical'], d['saha'], d['arizali'], d['kayip'],
                     d['unit'], d['description'], exists['id'])
                )
                updated_count += 1
            else:
                conn.execute(
                    """INSERT INTO depot_items
                       (category, name, current_stock, critical_stock, unit, description,
                        saha_stock, arizali_stock, kayip_stock)
                       VALUES (?,?,?,?,?,?,?,?,?)""",
                    (d['category'], d['name'], d['current'], d['critical'],
                     d['unit'], d['description'], d['saha'], d['arizali'], d['kayip'])
                )
                added_count += 1

        conn.commit()
        conn.close()
        print(f"DEBUG: Depot sync: {added_count} eklendi, {updated_count} güncellendi.")
        return jsonify({"message": f"Depo aktarıldı: {added_count} yeni ürün eklendi, {updated_count} mevcut ürün güncellendi."})
    except Exception as e:
        print(f"Depot sync error: {e}")
        return jsonify({"error": str(e)}), 500



@depot_manager_bp.route('/delete/<int:item_id>', methods=['DELETE'])
@require_admin
def delete_item(item_id):
    """Depo ürününü siler."""
    try:
        conn = get_db_connection()
        conn.execute("DELETE FROM depot_transactions WHERE depot_item_id=?", (item_id,))
        conn.execute("DELETE FROM depot_items WHERE id=?", (item_id,))
        conn.commit()
        conn.close()
        return jsonify({"message": "Ürün silindi"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@depot_manager_bp.route('/transaction', methods=['POST'])
@require_auth
def transaction():
    """Depo giriş/çıkış işlemi yapar. Cihaza atanıyorsa otomatik teknik not oluşturur."""
    data = request.json
    try:
        conn = get_db_connection()
        item_id = data.get('depot_item_id')
        t_type = data.get('type')  # 'in' veya 'out'
        qty = int(data.get('quantity', 0))

        if qty <= 0:
            return jsonify({"error": "Miktar 0'dan büyük olmalı"}), 400

        item = conn.execute("SELECT * FROM depot_items WHERE id=?", (item_id,)).fetchone()
        if not item:
            return jsonify({"error": "Ürün bulunamadı"}), 404

        current = item['current_stock']
        if t_type == 'out' and current < qty:
            return jsonify({"error": f"Yetersiz stok (Mevcut: {current})"}), 400

        new_stock = current + qty if t_type == 'in' else current - qty
        
        # Eğer bu bir SARF MALZEME veya Gıda ise ve işlem "çıkış" ise haftalık dağıtımı artır
        if t_type == 'out' and item['category'] in ['SARF MALZEME', 'Gıda']:
            conn.execute("UPDATE depot_items SET current_stock=?, weekly_distributed = weekly_distributed + ? WHERE id=?", (new_stock, qty, item_id))
        else:
            conn.execute("UPDATE depot_items SET current_stock=? WHERE id=?", (new_stock, item_id))

        conn.execute(
            "INSERT INTO depot_transactions (depot_item_id, transaction_type, quantity, device_id, device_type, user_name, note) VALUES (?,?,?,?,?,?,?)",
            (item_id, t_type, qty, data.get('device_id'), data.get('device_type'),
             data.get('user_name', 'System'), data.get('note', ''))
        )

        # Cihaza atanıyorsa otomatik teknik not oluştur
        if data.get('device_id') and t_type == 'out':
            conn.execute(
                "INSERT INTO technical_notes (device_id, device_type, title, content, user_id, user_name) VALUES (?,?,?,?,?,?)",
                (data.get('device_id'), data.get('device_type', 'pc'),
                 "Depodan Malzeme Atandı",
                 f"Depodan {qty} {item['unit']} {item['name']} ({item['category']}) bu cihaza atandı.",
                 data.get('user_id', 'system'), data.get('user_name', 'System'))
            )

        conn.commit()
        conn.close()
        return jsonify({"message": "İşlem tamamlandı", "new_stock": new_stock})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@depot_manager_bp.route('/history/<int:item_id>', methods=['GET'])
@require_auth
def get_history(item_id):
    """Bir depo ürününün tüm işlem geçmişini getirir."""
    try:
        items = query_db(
            "SELECT * FROM depot_transactions WHERE depot_item_id=? ORDER BY created_at DESC",
            (item_id,)
        )
        return jsonify([dict(row) for row in items])
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@depot_manager_bp.route('/clear_transactions', methods=['DELETE'])
@require_admin
def clear_all_transactions():
    """Tüm depo işlem geçmişini (depot_transactions) temizler."""
    try:
        conn = get_db_connection()
        result = conn.execute("DELETE FROM depot_transactions")
        deleted = result.rowcount
        conn.commit()
        conn.close()
        return jsonify({"message": f"{deleted} adet stok hareketi silindi.", "deleted": deleted})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
