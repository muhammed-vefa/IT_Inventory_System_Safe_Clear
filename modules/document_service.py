from flask import Blueprint, jsonify, send_from_directory, request, send_file
import os
import datetime
import shutil
import json
import openpyxl
import win32com.client
from fpdf import FPDF

document_service_bp = Blueprint('document_service', __name__)

BASE_DIR = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))
DL_DIR = os.path.join(BASE_DIR, "tools", "bat_uygulama")
TEMP_DIR = os.path.join(BASE_DIR, "uploads", "temp")

@document_service_bp.route('/list', methods=['GET'])
def list_files():
    try:
        if not os.path.exists(DL_DIR):
            os.makedirs(DL_DIR, exist_ok=True)
            
        files = []
        for f in os.listdir(DL_DIR):
            f_path = os.path.join(DL_DIR, f)
            if os.path.isfile(f_path):
                stats = os.stat(f_path)
                files.append({
                    "name": f,
                    "size": f"{round(stats.st_size / 1024, 2)} KB",
                    "date": datetime.datetime.fromtimestamp(stats.st_mtime).strftime("%Y-%m-%d %H:%M")
                })
        return jsonify({"success": True, "files": files})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@document_service_bp.route('/get/<filename>', methods=['GET'])
def get_file(filename):
    try:
        safe_filename = os.path.basename(filename)
        return send_from_directory(DL_DIR, safe_filename, as_attachment=True)
    except Exception as e:
        return jsonify({"success": False, "error": "Dosya bulunamadi"}), 404


# --- PDF GENERATION LOGIC ---

class TutanakPDF(FPDF):
    def __init__(self, t_type, **kwargs):
        super().__init__(**kwargs)
        self.t_type = t_type
        font_dir = os.path.join(os.path.dirname(__file__), "..", "static", "fonts")
        
        if os.path.exists(os.path.join(font_dir, "arial.ttf")):
            self.add_font("ArialTR", "", os.path.join(font_dir, "arial.ttf"), uni=True)
            self.add_font("ArialTR", "B", os.path.join(font_dir, "arialbd.ttf"), uni=True)
        else:
            self.add_font("ArialTR", "", "c:/windows/fonts/arial.ttf", uni=True)
            self.add_font("ArialTR", "B", "c:/windows/fonts/arialbd.ttf", uni=True)
        
    def header(self):
        if self.t_type == "VPN":
            return
        if self.t_type != "HT":
            LOGO_DIR = os.path.join(os.path.dirname(__file__), "..", "static", "logo")
            t_logo = "zimmet_left.png"
            t_path = os.path.join(LOGO_DIR, t_logo)
            if os.path.exists(t_path):
                self.image(t_path, 0, 0, 210)
        self.set_font("ArialTR", "B", 16)
        self.ln(25)
        self.ln(5)

    def footer(self):
        if self.t_type == "VPN":
            return
        if self.t_type != "HT":
            LOGO_DIR = os.path.join(os.path.dirname(__file__), "..", "static", "logo")
            b_logo = "zimmet_right.png"
            b_path = os.path.join(LOGO_DIR, b_logo)
            if os.path.exists(b_path):
                self.image(b_path, 0, 265, 210)

def generate_pdf_direct(t_type, items, photo_path=None):
    pdf = TutanakPDF(t_type)
    pdf.add_page()
    now_str = datetime.datetime.now().strftime("%d.%m.%Y")
    
    if t_type == "ZIMMET":
        pdf.set_y(35)
        pdf.set_font("ArialTR", "", 10)
        pdf.cell(0, 10, f"TARİH: {now_str}", ln=True, align="R")
        pdf.set_font("ArialTR", "B", 14)
        pdf.cell(0, 15, "ZİMMET TUTANAĞI", ln=True, align="C")
        pdf.ln(5)
        
        pdf.set_font("ArialTR", "", 10)
        staff = items.get("staff", "......")
        text = f"Aşağıda marka, model ve seri numaraları belirtilmiş cihazlar KEYDATA firmasından {staff} isimli personele / firmaya elden teslim edilmiştir."
        pdf.multi_cell(0, 6, text, align="C")
        pdf.ln(10)
        
        pdf.set_fill_color(245, 245, 245)
        pdf.set_font("ArialTR", "B", 10)
        pdf.set_x(10)
        pdf.cell(15, 8, "ADET", border=1, fill=True, align="C")
        pdf.cell(35, 8, "ÜRÜN TİPİ", border=1, fill=True, align="C")
        pdf.cell(40, 8, "MARKA", border=1, fill=True, align="C")
        pdf.cell(50, 8, "MODEL", border=1, fill=True, align="C")
        pdf.cell(50, 8, "SERİ NUMARASI", border=1, fill=True, align="C", ln=True)
        
        pdf.set_font("ArialTR", "", 10)
        for d in items.get("devices", []):
            pdf.set_x(10)
            pdf.cell(15, 8, str(d.get("adet", "1")), border=1, align="C")
            pdf.cell(35, 8, str(d.get("tip", "-")), border=1, align="C")
            pdf.cell(40, 8, str(d.get("marka", "-")), border=1, align="C")
            pdf.cell(50, 8, str(d.get("model", "-")), border=1, align="C")
            pdf.cell(50, 8, str(d.get("seri", "-")), border=1, align="C", ln=True)
            
        pdf.ln(10)
        pdf.set_x(10)
        pdf.cell(0, 6, "Durumu belirtilen iş bu tutanak tebellüğ yerine geçmesi hasebiyle imza altına alınmıştır.", ln=True)
        
        if pdf.get_y() > 240:
            pdf.add_page()
            
        pdf.ln(30)
        pdf.cell(95, 6, "Teslim Eden", align="C")
        pdf.cell(95, 6, "Teslim Alan", align="C", ln=True)
        pdf.cell(95, 6, "Ad-Soyad/Unvan", align="C")
        pdf.cell(95, 6, "Ad-Soyad/Unvan", align="C", ln=True)
        pdf.cell(95, 6, "İmza", align="C")
        pdf.cell(95, 6, "İmza", align="C", ln=True)
        pdf.ln(15)
        pdf.set_font("ArialTR", "B", 10)
        pdf.cell(95, 6, str(items.get("veren", "-")), align="C")
        pdf.cell(95, 6, str(items.get("alan", "-")), align="C", ln=True)
        pdf.set_font("ArialTR", "", 9)
        veren_unvan = items.get("veren_unvan", "Bilgi İşlem ve HBYS Uzm. Yrd.")
        alan_unvan = items.get("alan_unvan", "")
        pdf.cell(95, 6, veren_unvan, align="C")
        pdf.cell(95, 6, alan_unvan, align="C", ln=True)

    elif t_type == "SLA":
        pdf.set_auto_page_break(False)
        pdf.set_y(35)
        pdf.set_font("ArialTR", "B", 14)
        pdf.cell(0, 15, "TUTANAKTIR", ln=True, align="C")
        pdf.ln(10)
        
        pdf.set_font("ArialTR", "", 11)
        ticket = items.get("ticket", ".......")
        text1 = f"{ticket} no'lu SLA talebi sehven kapatılmıştır, açıklaması şu şekildedir;"
        pdf.multi_cell(0, 8, text1, align="L")
        pdf.ln(5)
        
        pdf.set_font("ArialTR", "B", 11)
        aciklama = items.get("aciklama", "Proje kapsamında yeteri sayıda bilgisayar ve ekipmanları bulunduğundan yeni bilgisayar, ekipman kurulumu yapılamayacaktır.")
        pdf.multi_cell(0, 8, aciklama, align="L")
        pdf.ln(5)
        
        pdf.set_font("ArialTR", "", 11)
        text3 = "İş bu tutanak bu açıklamalara istinaden tarafımızca imza altına toplanmıştır, gerekli işlemlerin yapılması konusunda destekleriniz rica olunur."
        pdf.multi_cell(0, 8, text3, align="L")
        
        pdf.ln(30)
        pdf.set_font("ArialTR", "", 9)
        kisi1_unvan = items.get("kisi1_unvan", "HBYS Yöneticisi")
        kisi2_unvan = items.get("kisi2_unvan", "HBYS Ve İYM Birim Sorumlusu")
        kisi3_unvan = items.get("kisi3_unvan", "Bilgi İşlem ve HBYS Uzm. Yrd.")
        pdf.cell(65, 6, kisi1_unvan, align="L")
        pdf.cell(65, 6, kisi2_unvan, align="C")
        pdf.cell(65, 6, kisi3_unvan, align="R", ln=True)
        
        pdf.set_font("ArialTR", "B", 10)
        kisi1_ad = items.get("kisi1_ad", "Ahmet Yılmaz")
        kisi2_ad = items.get("kisi2_ad", "Mehmet Demir")
        kisi3_ad = items.get("kisi3_ad", "Canan Yıldız")
        pdf.cell(65, 6, kisi1_ad, align="L")
        pdf.cell(65, 6, kisi2_ad, align="C")
        pdf.cell(65, 6, kisi3_ad, align="R", ln=True)

    elif t_type == "USULSUZ_TASIMA":
        pdf.set_auto_page_break(False)
        pdf.set_y(35)
        pdf.set_font("ArialTR", "B", 14)
        pdf.cell(0, 15, "TUTANAKTIR", ln=True, align="C")
        pdf.ln(10)
        
        pdf.set_font("ArialTR", "", 11)
        pc_no = items.get("pc_no", ".......")
        seri_no = items.get("seri_no", ".......")
        eski_mahal = items.get("eski_mahal", ".......")
        yeni_mahal = items.get("yeni_mahal", ".......")
        
        text1 = f"{pc_no} envanter ve {seri_no} seri numaralı donanımın, {eski_mahal} mahalinden {yeni_mahal} mahaline yetkisiz/onaysız olarak taşındığı tespit edilmiştir."
        pdf.multi_cell(0, 8, text1, align="L")
        pdf.ln(5)
        
        text3 = "İş bu tutanak tarafımızca imza altına toplanmıştır, gerekli işlemlerin yapılması konusunda destekleriniz rica olunur."
        pdf.multi_cell(0, 8, text3, align="L")
        
        pdf.ln(30)
        pdf.set_font("ArialTR", "", 9)
        kisi1_unvan = items.get("kisi1_unvan", "HBYS Yöneticisi")
        kisi2_unvan = items.get("kisi2_unvan", "HBYS Ve İYM Birim Sorumlusu")
        kisi3_unvan = items.get("kisi3_unvan", "Bilgi İşlem ve HBYS Uzm. Yrd.")
        pdf.cell(65, 6, kisi1_unvan, align="L")
        pdf.cell(65, 6, kisi2_unvan, align="C")
        pdf.cell(65, 6, kisi3_unvan, align="R", ln=True)
        
        pdf.set_font("ArialTR", "B", 10)
        kisi1_ad = items.get("kisi1_ad", "Ahmet Yılmaz")
        kisi2_ad = items.get("kisi2_ad", "Mehmet Demir")
        kisi3_ad = items.get("kisi3_ad", "")
        pdf.cell(65, 6, kisi1_ad, align="L")
        pdf.cell(65, 6, kisi2_ad, align="C")
        pdf.cell(65, 6, kisi3_ad, align="R", ln=True)

    elif t_type == "VPN":
        pdf.set_auto_page_break(False)
        # 1. HEADER
        pdf.set_line_width(0.5)
        pdf.rect(10, 10, 190, 30)
        pdf.line(55, 10, 55, 40)
        pdf.line(10, 35, 200, 35)
        pdf.line(55, 35, 55, 40)
        pdf.line(100, 35, 100, 40)
        pdf.line(140, 35, 140, 40)
        pdf.line(175, 35, 175, 40)
        
        logo_path = os.path.join(os.path.dirname(__file__), "..", "static", "logo", "ht_right.png")
        if os.path.exists(logo_path):
            pdf.image(logo_path, 25, 11, 14)
            
        pdf.set_font("ArialTR", "B", 7)
        pdf.set_xy(10, 26)
        pdf.cell(45, 3, "T.C. ÖRNEK BAKANLIĞI", align="C")
        pdf.set_xy(10, 29)
        pdf.set_font("ArialTR", "", 6)
        pdf.cell(45, 3, "ÖRNEKTEPE İL SAĞLIK MÜDÜRLÜĞÜ", align="C")
        pdf.set_xy(10, 32)
        pdf.cell(45, 3, "ÖRNEKTEPE DEVLET HASTANESİ", align="C")
        
        pdf.set_xy(55, 15)
        pdf.set_font("ArialTR", "B", 12)
        pdf.cell(145, 10, "ÖRNEKTEPE DEVLET HASTANESİ", align="C")
        pdf.set_xy(55, 22)
        pdf.cell(145, 10, "VPN BAĞLANTI TALEP FORMU", align="C")
        
        pdf.set_font("ArialTR", "B", 7)
        pdf.set_xy(10, 35)
        pdf.cell(45, 5, "DOKÜMAN KODU:BY.FR.12", align="C")
        pdf.set_xy(55, 35)
        pdf.cell(45, 5, "YAY.TAR.:01.11.2022", align="C")
        pdf.set_xy(100, 35)
        pdf.cell(40, 5, "REVİZYON TARİHİ: -", align="C")
        pdf.set_xy(140, 35)
        pdf.cell(35, 5, "REVİZYON NO: 00", align="C")
        pdf.set_xy(175, 35)
        pdf.cell(25, 5, "SAYFA 1 / 1", align="C")
        
        pdf.set_line_width(0.2)
        pdf.ln(10)
        
        # 2. Text Paragraph
        pdf.set_font("ArialTR", "", 11)
        text = (
            "Kampüs dışından kampüs ağına erişim için VPN (özel sanal ağ) hesabının açılmasını talep ediyorum. "
            "Açılacak VPN hesabı ile aşağıda belirtilen LAN (yerel ağ) bölgesindeki bilgisayara belirttiğim "
            "portlardan erişmek istiyorum. VPN hesabı ile erişim sağladığımda doğabilecek tüm sorumluluğun "
            "bende olduğunu, bağlantı istediğim sistemler dışında bir yere bağlanmayacağıma, kampüs ağ "
            "güvenliğine zarar vermeyeceğimi taahhüt ediyorum. Belirttiğim şartları sağlamadığım takdirde "
            "KEYDATA Bilişim Teknolojileri yetkililerinin bu hizmeti durdurabileceğini, inceleme ve yönetme "
            "konusunda yetkili olduğunu kabul ediyorum."
        )
        pdf.set_xy(10, 45)
        pdf.multi_cell(190, 5, text, align="J")
        pdf.ln(3)
        
        # 3. Kullanıcı Bilgileri
        pdf.set_fill_color(200, 200, 200)
        pdf.set_font("ArialTR", "B", 11)
        pdf.set_x(10)
        pdf.cell(190, 7, " Kullanıcı Bilgileri", border=1, fill=True, ln=True)
        
        pdf.set_font("ArialTR", "", 10)
        fields = [
            ("Adı ve Soyadı", items.get("adsoyad", "")),
            ("Firma Adı", items.get("firma", "")),
            ("Resmi Yazı Bilgisi", items.get("resmiyazi", "")),
            ("Görevi", items.get("gorevi", "")),
            ("HBYS Kul. Adı", items.get("hbys", "")),
            ("Cep Telefonu", items.get("telefon", "")),
            ("E-Posta", items.get("eposta", ""))
        ]
        
        for label, val in fields:
            pdf.set_x(10)
            pdf.cell(50, 7, f" {label}", border=1)
            pdf.cell(140, 7, f" {val}", border=1, ln=True)
            
        pdf.ln(3)
        
        # 4. OS Selection
        pdf.set_font("ArialTR", "B", 10)
        pdf.set_x(10)
        pdf.cell(190, 7, " VPN bağlantısını hangi işletim sisteminden yapmak istiyorsunuz?", border=1, fill=True, ln=True)
        
        pdf.set_font("ArialTR", "", 10)
        pdf.set_x(10)
        os_choice = items.get("os", "")
        pdf.cell(47.5, 7, f" Windows   {'[X]' if os_choice == 'Windows' else '[  ]'}", border=1)
        pdf.cell(47.5, 7, f" Linux        {'[X]' if os_choice == 'Linux' else '[  ]'}", border=1)
        pdf.cell(47.5, 7, f" Android     {'[X]' if os_choice == 'Android' else '[  ]'}", border=1)
        pdf.cell(47.5, 7, f" Mac IOS    {'[X]' if os_choice == 'Mac IOS' else '[  ]'}", border=1, ln=True)
        
        pdf.ln(3)
        
        # 5. VPN End Date
        pdf.set_x(10)
        pdf.set_font("ArialTR", "", 10)
        pdf.cell(190, 7, f"Vpn hesabının kapatılacağı tarih (sınırlı bir süre için ise) : .....{items.get('bitis', '')}.....", ln=True)
        pdf.ln(2)
        
        # 6. LAN Bilgileri
        pdf.set_x(10)
        pdf.set_font("ArialTR", "B", 10)
        pdf.cell(190, 6, "ERİŞİLMEK İSTENEN LAN (Yerel Ağ) BİLGİLERİ :", ln=True)
        
        pdf.set_font("ArialTR", "", 10)
        pdf.set_x(20)
        pdf.cell(180, 6, f"Network / Subnet (Var ise) : {items.get('network', '')}", ln=True)
        pdf.set_x(20)
        pdf.cell(180, 6, f"IP Adresi : {items.get('ip', '')}", ln=True)
        pdf.set_x(20)
        pdf.cell(180, 6, f"MAC (Ethernet) Adresi : {items.get('mac', '')}", ln=True)
        
        pdf.ln(5)
        
        # 7. Signatures
        pdf.set_x(10)
        pdf.cell(95, 6, "Yetkiyi İsteyen Adı Soyadı İmza (Kaşe)", align="C")
        pdf.cell(95, 6, "Hizmet Sağlayıcı Bilişim Teknolojileri", align="C", ln=True)
        pdf.set_x(10)
        pdf.cell(95, 6, "", align="C")
        pdf.cell(95, 6, "(Kaşe) İmza", align="C", ln=True)
        pdf.ln(8)
        pdf.set_x(10)
        pdf.cell(95, 6, "Tarih : ..../..../.......", align="C")
        pdf.cell(95, 6, "Tarih : ..../..../.......", align="C", ln=True)
        
        # 8. Footer table AT THE VERY BOTTOM OF THE PAGE
        pdf.set_y(-40)  # Pin to bottom
        pdf.set_font("ArialTR", "B", 10)
        pdf.set_line_width(0.5)
        
        pdf.set_x(20)
        pdf.cell(56.6, 8, "HAZIRLAYAN", border=1, align="C")
        pdf.cell(56.6, 8, "KONTROL EDEN", border=1, align="C")
        pdf.cell(56.6, 8, "ONAYLAYAN", border=1, align="C", ln=True)
        
        pdf.set_x(20)
        pdf.cell(56.6, 12, "", border=1, align="C")
        pdf.cell(56.6, 12, "KALİTE DİREKTÖRÜ", border=1, align="C")
        pdf.cell(56.6, 12, "BAŞHEKİM", border=1, align="C", ln=True)

    elif t_type == "HT":
        pdf.set_auto_page_break(False)
        
        # Outer border
        pdf.rect(5, 5, 200, 287)
        
        # 1. HEADER
        pdf.set_line_width(0.5)
        pdf.rect(10, 10, 190, 30)
        # Vertical lines for header
        pdf.line(55, 10, 55, 40)
        pdf.line(155, 10, 155, 40)
        # Horizontal lines for header bottom part
        pdf.line(10, 35, 200, 35)
        pdf.line(90, 35, 90, 40)
        pdf.line(125, 35, 125, 40)
        
        # Logos
        logo_left = os.path.join(os.path.dirname(__file__), "..", "static", "logo", "ht_left.png")
        if os.path.exists(logo_left):
            pdf.image(logo_left, 15, 18, 35)
            
        logo_right = os.path.join(os.path.dirname(__file__), "..", "static", "logo", "ht_right.png")
        if os.path.exists(logo_right):
            pdf.image(logo_right, 171, 11, 13)
            
        # Top right text
        pdf.set_font("ArialTR", "B", 6)
        pdf.set_xy(155, 25)
        pdf.cell(45, 3, "T.C. ÖRNEK BAKANLIĞI", align="C")
        pdf.set_xy(155, 28)
        pdf.set_font("ArialTR", "", 5)
        pdf.cell(45, 3, "ÖRNEKTEPE İL SAĞLIK MÜDÜRLÜĞÜ", align="C")
        pdf.set_xy(155, 31)
        pdf.cell(45, 3, "ÖRNEKTEPE DEVLET HASTANESİ", align="C")
        
        # Center title
        pdf.set_font("ArialTR", "B", 10)
        pdf.set_xy(55, 14)
        pdf.cell(100, 6, "ÖRNEKTEPE DEVLET HASTANESİ", align="C")
        pdf.set_font("ArialTR", "B", 14)
        pdf.set_xy(55, 22)
        pdf.cell(100, 8, "HASAR TESPİT TUTANAĞI", align="C")
        
        # Subheader texts
        pdf.set_font("ArialTR", "B", 8)
        pdf.set_xy(10, 35)
        pdf.cell(45, 5, "Doküman Kodu:MEF.MC.FR.30", align="C")
        pdf.set_xy(55, 35)
        pdf.cell(35, 5, "Yayın Tarihi:09.01.2019", align="C")
        pdf.set_xy(90, 35)
        pdf.cell(35, 5, "Revizyon No:0", align="C")
        pdf.set_xy(125, 35)
        pdf.cell(30, 5, "Revizyon Tarihi: -", align="C")
        pdf.set_xy(155, 35)
        pdf.cell(45, 5, "Sayfa:1/1", align="C")
        
        pdf.set_line_width(0.2)
        pdf.ln(10)
        
        # Main Title
        pdf.set_font("ArialTR", "B", 12)
        pdf.set_y(48)
        pdf.cell(0, 6, "BİLGİ İŞLEM HASAR TESPİT TUTANAĞI", ln=True, align="C")
        
        pdf.set_font("ArialTR", "", 10)
        pdf.set_y(58)
        now_str = datetime.datetime.now().strftime("%d.%m.%Y")
        pdf.cell(0, 6, f"Tarih : {now_str}", ln=True, align="R")
        pdf.set_y(65)
        sla = items.get("sla", "")
        pdf.cell(0, 6, f"SLA Numarası: {sla}", ln=True, align="L")
        
        pdf.ln(5)
        
        # 1. CİHAZ BİLGİLERİ
        pdf.set_font("ArialTR", "B", 10)
        pdf.cell(0, 6, "1. CİHAZ BİLGİLERİ", ln=True)
        pdf.set_font("ArialTR", "", 10)
        y_cihaz = pdf.get_y()
        pdf.set_fill_color(240, 240, 240)
        pdf.rect(10, y_cihaz, 190, 8)
        pdf.set_xy(12, y_cihaz + 1)
        model_str = items.get("model", "")
        seri_str = items.get("seri", items.get("serial", ""))
        pdf.cell(30, 6, "Ürün Modeli:", fill=True)
        pdf.cell(65, 6, f"{model_str}")
        pdf.cell(30, 6, "Seri Numarası:", fill=True)
        pdf.cell(60, 6, f"{seri_str}")
        pdf.ln(9)
        
        # 2. HASAR GÖREN CİHAZ TÜRÜ
        pdf.set_font("ArialTR", "B", 10)
        pdf.cell(0, 6, "2. HASAR GÖREN CİHAZ TÜRÜ", ln=True)
        y_start = pdf.get_y()
        pdf.rect(10, y_start, 190, 25)
        
        equipment = items.get("equipment", [])
        if isinstance(equipment, str):
            equipment = [equipment]
            
        def draw_checkbox(x, y, label, is_checked):
            pdf.rect(x, y, 3, 3)
            if is_checked:
                pdf.set_font("ArialTR", "B", 8)
                pdf.text(x+0.5, y+2.5, "X")
            pdf.set_font("ArialTR", "", 10)
            pdf.text(x+5, y+3, label)
            
        pdf.set_y(y_start + 4)
        draw_checkbox(15, pdf.get_y(), "Bilgisayar", "Bilgisayar" in equipment)
        draw_checkbox(65, pdf.get_y(), "Monitör", "Monitör" in equipment)
        draw_checkbox(115, pdf.get_y(), "Yazıcı", "Yazıcı" in equipment)
        draw_checkbox(160, pdf.get_y(), "Barkod Yazıcı", "Barkod Yazıcı" in equipment)
        pdf.ln(7)
        draw_checkbox(15, pdf.get_y(), "Barkod Okuyucu", "Barkod Okuyucu" in equipment)
        draw_checkbox(65, pdf.get_y(), "Switch", "Switch" in equipment)
        draw_checkbox(115, pdf.get_y(), "Klavye", "Klavye" in equipment)
        draw_checkbox(160, pdf.get_y(), "Mouse", "Mouse" in equipment)
        pdf.ln(7)
        draw_checkbox(15, pdf.get_y(), "43\" Ekran", "43\" Ekran" in equipment)
        draw_checkbox(65, pdf.get_y(), "24\" Ekran", "24\" Ekran" in equipment)
        draw_checkbox(115, pdf.get_y(), "Kiosk", "Kiosk" in equipment)
        other_val = items.get("other_equipment", "")
        draw_checkbox(160, pdf.get_y(), f"Diğer: {other_val}", "Diğer" in equipment or other_val != "")
        pdf.ln(10)
        
        # 3. HASAR TESPİT AÇIKLAMASI
        pdf.set_font("ArialTR", "B", 10)
        pdf.cell(0, 6, "3. HASAR TESPİT AÇIKLAMASI", ln=True)
        y_desc = pdf.get_y()
        pdf.rect(10, y_desc, 190, 35)
        pdf.set_xy(12, y_desc + 1)
        pdf.set_font("ArialTR", "", 8)
        pdf.set_text_color(120, 120, 120)
        pdf.cell(0, 4, "Hasar nasıl tespit edildi? Hasarın durumu ve kapsamı hakkında ayrıntılı bilgi veriniz.", ln=True)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("ArialTR", "", 10)
        desc = items.get("desc", "")
        pdf.set_xy(12, y_desc + 6)
        pdf.multi_cell(186, 6, desc, border=0)
        pdf.set_y(y_desc + 37)
        
        # 4. HASAR TESPİT FOTOĞRAFI
        pdf.set_font("ArialTR", "B", 10)
        pdf.cell(0, 6, "4. HASAR TESPİT FOTOĞRAFI", ln=True)
        photo_y = pdf.get_y()
        pdf.rect(10, photo_y, 190, 80)
        if photo_path and os.path.exists(photo_path):
            try:
                pdf.image(photo_path, 12, photo_y + 2, 186, 76)
            except Exception as e:
                pdf.set_xy(12, photo_y + 35)
                pdf.cell(186, 10, "Fotoğraf yüklenemedi", align="C")
        else:
            pdf.set_xy(12, photo_y + 35)
            pdf.set_font("ArialTR", "", 10)
            pdf.set_text_color(100, 100, 100)
            pdf.cell(186, 10, "(Lütfen hasarın net görüldüğü fotoğrafı buraya ekleyiniz veya zımbalayınız)", align="C")
            pdf.set_text_color(0, 0, 0)
            
        pdf.set_y(photo_y + 85)
        
        kullanici = items.get("teslimEden", items.get("kullanici", ""))
        kullanici_unvan = items.get("userUnvan", "")
        tespit_eden = items.get("tespitEden", items.get("tespit_eden", ""))
        tespit_unvan = items.get("tespitUnvan", "")
        birim_sorumlusu = items.get("birimSorumlusu", items.get("birim_sorumlusu", "Ahmet Yılmaz"))
        birim_unvan = items.get("birimUnvan", "")
        
        pdf.set_font("ArialTR", "", 9)
        pdf.cell(63, 6, "Kullanıcı / Sorumlu", align="C")
        pdf.cell(63, 6, "Tespit Eden", align="C")
        pdf.cell(64, 6, "Birim Sorumlusu", align="C", ln=True)
        
        pdf.cell(63, 6, "Ad-Soyad/Unvan/İmza", align="C")
        pdf.cell(63, 6, "Ad-Soyad/Unvan/İmza", align="C")
        pdf.cell(64, 6, "Ad-Soyad/Unvan/İmza", align="C", ln=True)
        
        pdf.set_font("ArialTR", "", 9)
        pdf.cell(63, 6, kullanici_unvan, align="C")
        pdf.cell(63, 6, tespit_unvan, align="C")
        pdf.cell(64, 6, birim_unvan, align="C", ln=True)
        
        pdf.set_font("ArialTR", "B", 9)
        pdf.cell(63, 6, kullanici, align="C")
        pdf.cell(63, 6, tespit_eden, align="C")
        pdf.cell(64, 6, birim_sorumlusu, align="C", ln=True)

    temp_dir = TEMP_DIR
    os.makedirs(temp_dir, exist_ok=True)
    temp_pdf_path = os.path.join(temp_dir, f"temp_direct_{datetime.datetime.now().timestamp()}.pdf")
    pdf.output(temp_pdf_path)
    return temp_pdf_path

def generate_ht_from_excel(items, photo_path=None):
    temp_dir = TEMP_DIR
    os.makedirs(temp_dir, exist_ok=True)
    
    timestamp = datetime.datetime.now().timestamp()
    temp_excel = os.path.join(temp_dir, f"temp_ht_{timestamp}.xlsx")
    temp_pdf = os.path.join(temp_dir, f"temp_ht_{timestamp}.pdf")
    
    template_path = os.path.join(os.path.dirname(__file__), "..", "database", "sablonlar", "hasar_tespit.xlsx")
    shutil.copy2(template_path, temp_excel)
    
    wb = openpyxl.load_workbook(temp_excel)
    ws = wb.active
    
    # Text Fields
    ws["D10"] = str(items.get("sla", "-"))
    ws["D32"] = str(items.get("seri", "-"))
    ws["D33"] = str(items.get("model", "-"))
    ws["B27"] = str(items.get("desc", items.get("hasar_aciklama", "-")))
    ws["B50"] = str(items.get("teslimEden", "-"))
    ws["E50"] = str(items.get("tespitEden", "-"))
    ws["I50"] = str(items.get("birimSorumlusu", "-"))
    
    # Checkboxes
    equipment_cells = {
        "Bilgisayar": "E12", "Klavye": "I12", "Monitör": "E14", "Mouse": "I14",
        "Yazıcı": "E16", "43\" Ekran": "I16", "Barkod Yazıcı": "E18", "24\" Ekran": "I18",
        "Barkod Okuyucu": "E20", "Kiosk": "I20", "Switch": "I22"
    }
    
    selected_eqs = items.get("equipment", [])
    if isinstance(selected_eqs, str):
        selected_eqs = [selected_eqs]
        
    other_items = []
    for eq in selected_eqs:
        if eq in equipment_cells:
            ws[equipment_cells[eq]] = "X"
        else:
            other_items.append(eq)
            
    if other_items:
        ws["E22"] = "X"
        ws["D22"] = ", ".join(other_items)
        
    # Photo insertion
    if photo_path and os.path.exists(photo_path):
        try:
            from openpyxl.drawing.image import Image as OpenpyxlImage
            img = OpenpyxlImage(photo_path)
            # Resize image to fit nicely
            img.width = 320
            img.height = 160
            ws.add_image(img, "G35")
        except Exception as e:
            print("Could not add image to Excel:", e)

    wb.save(temp_excel)
    
    # Convert to PDF
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        wb_com = excel.Workbooks.Open(os.path.abspath(temp_excel))
        ws_com = wb_com.ActiveSheet
        ws_com.PageSetup.Zoom = False
        ws_com.PageSetup.FitToPagesWide = 1
        ws_com.PageSetup.FitToPagesTall = False
        ws_com.PageSetup.LeftMargin = 5
        ws_com.PageSetup.RightMargin = 5
        ws_com.PageSetup.TopMargin = 5
        ws_com.PageSetup.BottomMargin = 5
        
        wb_com.ExportAsFixedFormat(0, os.path.abspath(temp_pdf))
        wb_com.Close(False)
        excel.Quit()
    except Exception as e:
        print("Excel to PDF conversion failed:", e)
        # Fallback to excel if PDF fails
        return temp_excel, [temp_excel]
        
    return temp_pdf, [temp_excel, temp_pdf]

def generate_izin_from_excel(items):
    temp_dir = TEMP_DIR
    os.makedirs(temp_dir, exist_ok=True)
    
    timestamp = datetime.datetime.now().timestamp()
    temp_excel = os.path.join(temp_dir, f"temp_izin_{timestamp}.xlsx")
    temp_pdf = os.path.join(temp_dir, f"temp_izin_{timestamp}.pdf")
    
    template_path = os.path.join(os.path.dirname(__file__), "..", "database", "sablonlar", "İzin İstek Formu.xlsx")
    shutil.copy2(template_path, temp_excel)
    
    def format_date(d_str):
        if not d_str or d_str == "-":
            return "-"
        try:
            parts = d_str.split("-")
            if len(parts) == 3 and len(parts[0]) == 4:
                return f"{parts[2]}.{parts[1]}.{parts[0]}"
        except:
            pass
        return d_str

    wb = openpyxl.load_workbook(temp_excel)
    ws = wb.active
    
    # Fill values
    ws["B6"] = datetime.datetime.now().strftime("%d.%m.%Y")
    ws["B7"] = str(items.get("ad_soyad", "-"))
    ws["F7"] = str(items.get("bolum", "-"))
    ws["B8"] = str(items.get("sicil", "-"))
    ws["F8"] = str(items.get("gorev", "-"))
    ws["B9"] = str(items.get("sebep", "-"))
    ws["B10"] = format_date(items.get("baslangic", "-"))
    ws["D10"] = str(items.get("bas_saat", "-"))
    ws["B11"] = format_date(items.get("bitis", "-"))
    ws["D11"] = str(items.get("bit_saat", "-"))
    ws["B12"] = format_date(items.get("isbasi", "-"))
    ws["D12"] = str(items.get("isbasi_saat", "-"))
    
    # Format duration
    ws["E10"] = f"{items.get('sure_gun', '0')} Gün / {items.get('sure_saat', '0')} Saat"
    
    # Paid/Unpaid selection in A13
    tur = items.get("tur", "")
    if "Ücretsiz" in tur:
        ws["A13"] = "Yukarıda Adı Soyadı Yazılı çalışanımıza mazeretine binaen aşağıda belirtilen tarih / tarihleri arasında ÜCRETSİZ izin verilmesi uygun görülmüştür."
    else:
        ws["A13"] = "Yukarıda Adı Soyadı Yazılı çalışanımıza mazeretine binaen aşağıda belirtilen tarih / tarihleri arasında ÜCRETLİ izin verilmesi uygun görülmüştür."
        
    # Signatures
    ws["A16"] = str(items.get("talep_eden_ad", "-"))
    ws["B16"] = str(items.get("takim_lideri_ad", "-"))
    ws["C16"] = str(items.get("bolum_muduru_ad", "-"))
    ws["F16"] = str(items.get("ik_ad", "-"))
    
    wb.save(temp_excel)
    
    # Convert to PDF
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        wb_com = excel.Workbooks.Open(os.path.abspath(temp_excel))
        ws_com = wb_com.ActiveSheet
        ws_com.PageSetup.Zoom = False
        ws_com.PageSetup.FitToPagesWide = 1
        ws_com.PageSetup.FitToPagesTall = False
        ws_com.PageSetup.LeftMargin = 5
        ws_com.PageSetup.RightMargin = 5
        ws_com.PageSetup.TopMargin = 5
        ws_com.PageSetup.BottomMargin = 5
        
        wb_com.ExportAsFixedFormat(0, os.path.abspath(temp_pdf))
        wb_com.Close(False)
        excel.Quit()
    except Exception as e:
        print("Excel to PDF conversion failed for Izin:", e)
        # Fallback to excel if PDF fails
        return temp_excel, [temp_excel]
        
    return temp_pdf, [temp_excel, temp_pdf]

@document_service_bp.route('/generate_tutanak', methods=['POST'])
def generate_tutanak():
    files_to_delete = []
    try:
        temp_dir = TEMP_DIR
        os.makedirs(temp_dir, exist_ok=True)
        
        photo_path = None
        if request.is_json:
            data = request.json
            items = data.get('data', data.get('items', {}))
            t_type = data.get('type')
        else:
            data = request.form
            items = json.loads(data.get('data', '{}'))
            t_type = data.get('type')
            photo_file = request.files.get('photo')
            if photo_file:
                photo_path = os.path.join(temp_dir, f"photo_{datetime.datetime.now().timestamp()}.jpg")
                photo_file.save(photo_path)
                files_to_delete.append(photo_path)
        if not t_type:
            return jsonify({"success": False, "error": "type is required"}), 400
            
        if t_type == "IZIN":
            out_path, created_files = generate_izin_from_excel(items)
            files_to_delete.extend(created_files)
            response = send_file(out_path, as_attachment=True, download_name=f"{t_type}_Tutanak.{out_path.split('.')[-1]}")
        else:
            pdf_path = generate_pdf_direct(t_type, items, photo_path)
            files_to_delete.append(pdf_path)
            response = send_file(pdf_path, as_attachment=True, download_name=f"{t_type}_Tutanak.pdf")

        @response.call_on_close
        def remove_temporary_files():
            import time
            time.sleep(0.5)  # Buffer for Win32 Excel/Flask file handles to close
            for path in files_to_delete:
                last_err = None
                for _ in range(3):
                    try:
                        if path and os.path.exists(path):
                            os.remove(path)
                        break
                    except Exception as e:
                        last_err = e
                        time.sleep(0.2)
                else:
                    print(f"Error removing temporary file {path}: {last_err}")

        return response
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
