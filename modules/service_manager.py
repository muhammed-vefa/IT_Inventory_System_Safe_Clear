from core.utils import normalize_row
from flask import Blueprint, jsonify, request
from core.database_sql import get_db_connection
from core.auth import require_auth, require_admin, require_editor

service_manager_bp = Blueprint('service_manager', __name__)

def extract_printer_number(name):
    if not name:
        return None
    digits = "".join(filter(str.isdigit, str(name)))
    if digits:
        return int(digits)
    return None

@service_manager_bp.route('/get_all', methods=['GET'])
@require_auth
def get_all():
    import traceback
    try:
        conn = get_db_connection()
        if not conn: return jsonify([])
        cursor = conn.cursor()
        query = """
            SELECT s.*, p.location_code as printer_location 
            FROM printer_service s
            LEFT JOIN printers p ON s.pr_no = p.pr_no
            WHERE s.is_deleted = 0
            ORDER BY CASE WHEN s.acquisition_date IS NULL THEN 0 ELSE 1 END DESC, s.acquisition_date DESC, s.id DESC
        """
        cursor.execute(query)
        columns = [column[0].lower() for column in cursor.description]
        
        results = []
        for row in cursor.fetchall():
            d = dict(zip(columns, row))
            if not d.get('location_code'):
                d['location_code'] = d.get('printer_location')
                
            # Servis tablosunda "SERVİSTE-" veya "DEPO-" önekini gizle
            raw_mahal = str(d.get('location_code') or '')
            if raw_mahal.startswith('SERVİSTE-'):
                raw_mahal = raw_mahal.replace('SERVİSTE-', '', 1)
            elif raw_mahal.startswith('DEPO-'):
                raw_mahal = raw_mahal.replace('DEPO-', '', 1)
                
            d['location_code'] = raw_mahal
            d['mahal'] = raw_mahal.replace('-', '.') if raw_mahal else ''
            
            # Frontend seri bekliyor
            d['seri'] = d.get('serial_no')
            
            # Eğer geçmiş kayıtlarda status NULL ise, tarihlere bakarak toparla
            if not d.get('status'):
                if d.get('return_date'):
                    d['status'] = 'Tamamlandı'
                elif d.get('sent_date'):
                    d['status'] = 'Serviste'
                else:
                    d['status'] = 'Arızalı'
                    
            results.append(normalize_row(d))
            
        conn.close()
        return jsonify(results)
    except Exception as e:
        print(f"[get_all ERROR] {e}")
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

from datetime import datetime

def parse_date_safely(val):
    if not val:
        return None
    val_str = str(val).strip()
    if val_str.lower() in ('', 'none', 'null', '-', 'undefined'):
        return None
    
    formats = [
        "%d.%m.%Y",          # 18.05.2026
        "%Y-%m-%d",          # 2026-05-18
        "%Y-%m-%d %H:%M:%S", # 2026-05-18 16:30:00
        "%d/%m/%Y",          # 18/05/2026
        "%Y/%m/%d"           # 2026/05/18
    ]
    for fmt in formats:
        try:
            return datetime.strptime(val_str, fmt)
        except ValueError:
            continue
    return None


def find_printer_row(cursor, pr_no):
    if not pr_no:
        return None
    cursor.execute("SELECT id, pr_no, location_code FROM printers WHERE pr_no=? AND is_deleted=0", (pr_no,))
    row = cursor.fetchone()
    if row:
        return row
    
    digits = "".join(filter(str.isdigit, str(pr_no)))
    if digits:
        formatted_pr_no = f"PR-{digits.zfill(3)}"
        cursor.execute("SELECT id, pr_no, location_code FROM printers WHERE (pr_no=? OR pr_no=?) AND is_deleted=0", (formatted_pr_no, digits))
        row = cursor.fetchone()
        if row:
            return row
    return None

def swap_connected_printer_string(connected, old_no, new_no):
    if not connected:
        return connected
    
    old_clean = "".join(filter(str.isdigit, str(old_no)))
    new_clean = "".join(filter(str.isdigit, str(new_no)))
    if not old_clean or not new_clean:
        return connected
    
    old_pr = f"PR-{old_clean.zfill(3)}"
    old_pr_lower = old_pr.lower()
    
    new_pr = f"PR-{new_clean.zfill(3)}"
    new_pr_lower = new_pr.lower()
    
    res = connected
    res = res.replace(old_pr, new_pr)
    res = res.replace(old_pr_lower, new_pr_lower)
    
    import re
    res = re.sub(r'\b' + re.escape(old_clean) + r'\b', new_clean, res)
    res = re.sub(r'\b' + re.escape(old_clean.lstrip('0')) + r'\b', new_clean.lstrip('0'), res)
    return res

def run_service_automations_async(user_id, pr_no, substitute_pr_no, orig_location, pc_targets, has_substitute, status_str, is_return=False):
    import threading
    
    def worker():
        try:
            from core.database_sql import get_db_connection
            from core.encryption import decrypt_password
            from modules.inventory_printers import update_cups_printer_location_wizard, cups_do_pause_reject, cups_do_resume_accept
            import requests
            import urllib.parse

            print(f"[Async Automation] Started. user_id: {user_id}, pr_no: {pr_no}, sub_pr_no: {substitute_pr_no}, is_return: {is_return}")

            if is_return:
                # --- RETURN FROM SERVICE FLOW ---
                # The returned printer (pr_no) ALWAYS goes to Kontrolde first to be checked, and is paused/rejected.
                # The substitute printer (substitute_pr_no) stays untouched in its new location.
                update_cups_printer_location_wizard(pr_no, "Kontrolde")
                # Pause the returning printer
                succ, msg, _ = cups_do_pause_reject(pr_no)
                print(f"[Async CUPS Return to Depo] {pr_no}: {msg}")

            else:
                # --- SEND TO SERVICE FLOW ---
                # 1. CUPS Faulty Printer location change & pause/reject
                if has_substitute:
                    faulty_cups_location = "SERVİSTE"
                else:
                    faulty_cups_location = f"SERVİSTE-{orig_location}" if orig_location and not orig_location.startswith('SERVİSTE') else "SERVİSTE"
                update_cups_printer_location_wizard(pr_no, faulty_cups_location)
                
                succ, msg, _ = cups_do_pause_reject(pr_no)
                print(f"[Async CUPS Pause/Reject] {pr_no}: {msg}")

                # 2. CUPS Substitute Printer location update & resume/accept
                if has_substitute and substitute_pr_no and orig_location:
                    update_cups_printer_location_wizard(substitute_pr_no, orig_location)
                    succ, msg, _ = cups_do_resume_accept(substitute_pr_no)
                    print(f"[Async CUPS Substitute Resume/Accept] {substitute_pr_no}: {msg}")

            # 3. BIM PC Mappings Update
            if not pc_targets:
                print("[Async BIM] No target PCs to update. Done.")
                return

            conn = get_db_connection()
            if not conn:
                print("[Async BIM Error] Database connection failed")
                return
            cursor = conn.cursor()
            cursor.execute("SELECT bim_user, bim_pass FROM users WHERE id = ?", (user_id,))
            user_row = cursor.fetchone()
            conn.close()

            if not user_row or not user_row[0] or not user_row[1]:
                print("[Async BIM Warning] No BIM credentials found for user. Skipping PC automation.")
                return

            bim_user = user_row[0]
            try:
                bim_pass = decrypt_password(user_row[1])
            except Exception as dec_err:
                print(f"[Async BIM Error] Decrypt pass error: {dec_err}")
                return

            base_url = "http://bim.kocaelish.com/Handler.ashx"
            
            for pc in pc_targets:
                pc_ip = pc.get('ip')
                pc_name = pc.get('pc_no')
                if not pc_ip:
                    continue

                bim_session = requests.Session()
                browser_headers = {
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
                    "X-Requested-With": "XMLHttpRequest",
                    "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
                    "Client-IP": pc_ip,
                    "X-Forwarded-For": pc_ip,
                    "Origin": "http://bim.kocaelish.com",
                    "Referer": "http://bim.kocaelish.com/"
                }

                # BIM Login
                try:
                    login_data = {
                        "Functions": "Login",
                        "UserName": bim_user,
                        "Password": bim_pass
                    }
                    encoded_login = urllib.parse.urlencode(login_data)
                    login_resp = bim_session.post(base_url, data=encoded_login, headers=browser_headers, timeout=10, verify=False)
                    if login_resp.status_code != 200 or login_resp.text.strip() == "Error" or not login_resp.text.strip():
                        print(f"[Async BIM Error] Login failed for PC {pc_name}")
                        continue
                    ipa_session = login_resp.text.strip()
                except Exception as le:
                    print(f"[Async BIM Error] Login exception for PC {pc_name}: {le}")
                    continue

                headers = browser_headers.copy()
                headers["IPASession"] = ipa_session

                if is_return:
                    # Remove substitute printer from client PC
                    if substitute_pr_no:
                        try:
                            rem_data = {
                                "UserName": bim_user,
                                "IPAddress": pc_ip,
                                "Functions": "RemovePrinter",
                                "PrinterName": substitute_pr_no
                            }
                            enc_rem = urllib.parse.urlencode(rem_data)
                            rem_resp = bim_session.post(base_url, data=enc_rem, headers=headers, timeout=15, verify=False)
                            print(f"[Async BIM] RemovePrinter {substitute_pr_no} from {pc_ip} - Status: {rem_resp.status_code}, Body: {rem_resp.text.strip()}")
                        except Exception as re_err:
                            print(f"[Async BIM Error] RemovePrinter failed for PC {pc_name}: {re_err}")
                    
                    # Add repaired printer back to client PC
                    try:
                        add_data = {
                            "UserName": bim_user,
                            "IPAddress": pc_ip,
                            "Functions": "AddPrinter",
                            "PrinterName": f"{pr_no}/01"
                        }
                        enc_add = urllib.parse.urlencode(add_data)
                        add_resp = bim_session.post(base_url, data=enc_add, headers=headers, timeout=15, verify=False)
                        print(f"[Async BIM] AddPrinter {pr_no}/01 to {pc_ip} - Status: {add_resp.status_code}, Body: {add_resp.text.strip()}")
                    except Exception as ad_err:
                        print(f"[Async BIM Error] AddPrinter failed for PC {pc_name}: {ad_err}")

                else:
                    # Remove faulty printer from client PC
                    try:
                        rem_data = {
                            "UserName": bim_user,
                            "IPAddress": pc_ip,
                            "Functions": "RemovePrinter",
                            "PrinterName": pr_no
                        }
                        enc_rem = urllib.parse.urlencode(rem_data)
                        rem_resp = bim_session.post(base_url, data=enc_rem, headers=headers, timeout=15, verify=False)
                        print(f"[Async BIM] RemovePrinter {pr_no} from {pc_ip} - Status: {rem_resp.status_code}, Body: {rem_resp.text.strip()}")
                    except Exception as re_err:
                        print(f"[Async BIM Error] RemovePrinter failed for PC {pc_name}: {re_err}")

                    # Add substitute printer to client PC
                    if has_substitute and substitute_pr_no:
                        try:
                            add_data = {
                                "UserName": bim_user,
                                "IPAddress": pc_ip,
                                "Functions": "AddPrinter",
                                "PrinterName": f"{substitute_pr_no}/01"
                            }
                            enc_add = urllib.parse.urlencode(add_data)
                            add_resp = bim_session.post(base_url, data=enc_add, headers=headers, timeout=15, verify=False)
                            print(f"[Async BIM] AddPrinter {substitute_pr_no}/01 to {pc_ip} - Status: {add_resp.status_code}, Body: {add_resp.text.strip()}")
                        except Exception as ad_err:
                            print(f"[Async BIM Error] AddPrinter failed for PC {pc_name}: {ad_err}")

            print("[Async Automation] Finished successfully.")

        except Exception as general_err:
            print(f"[Async Automation Worker Error] {general_err}")

    t = threading.Thread(target=worker)
    t.daemon = True
    t.start()


@service_manager_bp.route('/add', methods=['POST'])
@require_auth
def add_service():
    try:
        data = request.json
        print("ADD SERVICE PAYLOAD:", data)
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # UI'dan gelen tarihleri kontrol et, bos ise None yap ve guvenli parse et
        acq_date = parse_date_safely(data.get('acquisition_date') or data.get('acq_date'))
        sent_date = parse_date_safely(data.get('sent_date'))
        return_date = parse_date_safely(data.get('return_date'))
        fault_desc = data.get('fault_description') or data.get('fault_desc') or None
        
        if fault_desc and str(fault_desc).strip() in ('', 'None', 'null', 'undefined'):
            fault_desc = None

        if return_date:
            data['status'] = 'Tamamlandı'
            p_status = (0, 0, 1, 0) # is_faulty, in_service, warehouse, on_field
        elif sent_date:
            data['status'] = 'Serviste'
            p_status = (0, 1, 0, 0)
        else:
            data['status'] = 'Arızalı'
            p_status = (1, 0, 0, 0)

        print("ADD PARSED DATES - acq_date:", acq_date, "sent_date:", sent_date, "return_date:", return_date)

        pr_no = data.get('pr_no')
        substitute_pr_no = data.get('substitute_pr_no')
        has_substitute = 1 if data.get('has_substitute') else 0

        # Find faulty printer row and normalize pr_no
        faulty_row = find_printer_row(cursor, pr_no)
        original_location = None
        if faulty_row:
            pr_no = faulty_row[1] # Normalized pr_no
            original_location = faulty_row[2] # location_code

        raw_location = original_location or data.get('mahal') or data.get('location_code') or 'BİLİNMİYOR'
        current_location = raw_location.replace('SERVİSTE-', '').replace('DEPO-', '').strip() or 'BİLİNMİYOR'

        # Insert service record
        cursor.execute("""
            INSERT INTO printer_service (
                pr_no, sla_no, serial_no, mac, model, fault_description, 
                status, acquisition_date, sent_date, return_date, 
                has_substitute, substitute_pr_no, user_name, location_code
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            pr_no, data.get('sla_no'), data.get('seri') or data.get('serial_no'), 
            data.get('mac'), data.get('model'), 
            fault_desc, data.get('status'), 
            acq_date, sent_date, return_date,
            has_substitute,
            substitute_pr_no,
            data.get('user_name', 'system'),
            current_location
        ))

        # PC connected printers updates
        pc_targets = []
        if pr_no:
            old_digits = "".join(filter(str.isdigit, str(pr_no)))
            cursor.execute("SELECT id, pc_no, ip, connected_printers FROM pcs WHERE is_deleted=0 AND (connected_printers LIKE ? OR connected_printers LIKE ? OR connected_printers LIKE ?)", 
                           (f"%{pr_no}%", f"%PR-{old_digits.zfill(3)}%", f"%{old_digits}%"))
            pc_rows = cursor.fetchall()
            for pc_id, pc_name, pc_ip, connected in pc_rows:
                pc_targets.append({
                    "id": pc_id,
                    "pc_no": pc_name,
                    "ip": pc_ip
                })
                if has_substitute and substitute_pr_no:
                    new_conn_str = swap_connected_printer_string(connected, pr_no, substitute_pr_no)
                    cursor.execute("UPDATE pcs SET connected_printers = ? WHERE id = ?", (new_conn_str, pc_id))

        # Update faulty printer in DB (always set location_code and location to have correct display everywhere)
        if return_date:
            # Servisten dönen yazıcı kontrol edilene kadar KONTROLDE bekler
            clean_loc = current_location.replace('SERVİSTE-', '').replace('DEPO-', '').replace('ARIZALI-', '').replace('KONTROLDE-', '').strip() or 'BİLİNMİYOR'
            if has_substitute:
                kontrol_location = "KONTROLDE"
            else:
                kontrol_location = f"KONTROLDE-{clean_loc}"
            cursor.execute("UPDATE printers SET is_faulty=0, in_service=0, warehouse=0, on_field=0, location_code=? WHERE pr_no=?", (kontrol_location, pr_no))
            
            # Since we inserted it directly as returned, mark it as uncontrolled (requires checking first)
            try:
                cursor.execute("ALTER TABLE printer_service ADD is_controlled BIT DEFAULT 0")
            except:
                pass
            cursor.execute("UPDATE printer_service SET is_controlled=0 WHERE pr_no=? AND return_date IS NOT NULL", (pr_no,))
        else:
            loc = current_location
            if has_substitute:
                # İkame varsa arızalı yazıcı direkt SERVİSTE olur, kendi mahalini kaybeder.
                new_location = "SERVİSTE"
            else:
                # İkame yoksa arızalı yazıcı SERVİSTE-Mahal olur.
                clean_loc = loc.replace('SERVİSTE-', '').replace('DEPO-', '').strip() or loc
                new_location = f"SERVİSTE-{clean_loc}" if not loc.startswith('SERVİSTE-') and loc != 'SERVİSTE' else loc
            
            cursor.execute("UPDATE printers SET is_faulty=?, in_service=?, warehouse=?, on_field=?, location_code=? WHERE pr_no=?", (*p_status, new_location, pr_no))

        # Update substitute printer in DB
        if has_substitute and substitute_pr_no:
            sub_row = find_printer_row(cursor, substitute_pr_no)
            if sub_row:
                substitute_pr_no = sub_row[1] # Normalized substitute number
            # İkame yazıcı arızalı yazıcının mahalini alır ve Kurulu (on_field=1) olur.
            cursor.execute("UPDATE printers SET is_faulty=0, in_service=0, warehouse=0, on_field=1, location_code=? WHERE pr_no=?", (current_location, substitute_pr_no))
            cursor.execute("SELECT id FROM printers WHERE (pr_no=? OR pr_no=?) AND is_deleted=0", (substitute_pr_no, "".join(filter(str.isdigit, str(substitute_pr_no)))))
            sub_row = cursor.fetchone()
            if sub_row:
                cursor.execute("UPDATE printers SET is_faulty=0, in_service=0, warehouse=0, on_field=1, location_code=? WHERE id=?", 
                               (original_location or current_location, sub_row[0]))

        conn.commit()
        conn.close()

        # START BACKGROUND AUTOMATION
        current_user = getattr(request, 'current_user', {})
        uid = current_user.get('user_id')
        
        run_service_automations_async(
            user_id=uid,
            pr_no=pr_no,
            substitute_pr_no=substitute_pr_no,
            orig_location=original_location,
            pc_targets=pc_targets,
            has_substitute=has_substitute,
            status_str=data['status'],
            is_return=bool(return_date)
        )

        return jsonify({"success": True})
    except Exception as e:
        import traceback
        print(f"[add_service ERROR] {e}")
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@service_manager_bp.route('/update/<int:record_id>', methods=['PUT'])
@require_auth
def update_service(record_id):
    try:
        user_role = request.current_user.get('role', '')
        if user_role not in ('ADMIN', 'DEPOT'):
            return jsonify({"success": False, "error": "Bu işlem için yetkiniz bulunmamaktadır. Sadece Admin ve Depocu işlem yapabilir."}), 403

        data = request.json
        conn = get_db_connection()
        if not conn:
            return jsonify({"success": False, "error": "Database connection failed"}), 500
        cursor = conn.cursor()

        # Check existing record in database
        cursor.execute("SELECT return_date, pr_no, substitute_pr_no, has_substitute, location_code FROM printer_service WHERE id = ?", (record_id,))
        existing_row = cursor.fetchone()
        if not existing_row:
            conn.close()
            return jsonify({"success": False, "error": "Servis kaydı bulunamadı."}), 404
            
        existing_return_date, old_pr_no, old_substitute_pr_no, old_has_substitute, existing_location_code = existing_row
        if existing_return_date is not None and user_role != 'ADMIN':
            conn.close()
            return jsonify({"success": False, "error": "Geldiği tarih bilgisi girilmiş kapalı kayıtları sadece Admin düzenleyebilir."}), 403
        
        acq_date = parse_date_safely(data.get('acquisition_date') or data.get('acq_date'))
        sent_date = parse_date_safely(data.get('sent_date'))
        return_date = parse_date_safely(data.get('return_date'))
        fault_desc = data.get('fault_description') or data.get('fault_desc') or None
        
        if fault_desc and str(fault_desc).strip() in ('', 'None', 'null', 'undefined'):
            fault_desc = None

        if return_date:
            data['status'] = 'Tamamlandı'
            p_status = (0, 0, 1, 0) # is_faulty, in_service, warehouse, on_field
        elif sent_date:
            data['status'] = 'Serviste'
            p_status = (0, 1, 0, 0)
        else:
            data['status'] = 'Arızalı'
            p_status = (1, 0, 0, 0)

        pr_no = data.get('pr_no') or old_pr_no
        substitute_pr_no = data.get('substitute_pr_no') or old_substitute_pr_no
        has_substitute = 1 if data.get('has_substitute') else 0

        # Retrieve printer row to normalize pr_no
        pr_row = find_printer_row(cursor, pr_no)
        if pr_row:
            pr_no = pr_row[1]

        # Use the provided mahal from the frontend (which user might have edited).
        # Fallback to the existing location code if not provided.
        # Clean DEPO/SERVİSTE prefixes just in case.
        provided_mahal = data.get('mahal')
        if provided_mahal and str(provided_mahal).strip():
            current_location = str(provided_mahal).replace('SERVİSTE-', '').replace('DEPO-', '').strip()
        else:
            current_location = existing_location_code or 'BİLİNMİYOR'

        # DETECT IF PRINTER IS RETURNING NOW
        is_returning = (existing_return_date is None and return_date is not None)

        # PCs connected printers updates
        pc_targets = []
        if not is_returning:
            # Normal check/update PC mappings if substitute exists (Only on send to service or update)
            if has_substitute and substitute_pr_no:
                old_digits = "".join(filter(str.isdigit, str(pr_no)))
                cursor.execute("SELECT id, pc_no, ip, connected_printers FROM pcs WHERE is_deleted=0 AND (connected_printers LIKE ? OR connected_printers LIKE ? OR connected_printers LIKE ?)", 
                               (f"%{pr_no}%", f"%PR-{old_digits.zfill(3)}%", f"%{old_digits}%"))
                pc_rows = cursor.fetchall()
                for pc_id, pc_name, pc_ip, connected in pc_rows:
                    pc_targets.append({
                        "id": pc_id,
                        "pc_no": pc_name,
                        "ip": pc_ip
                    })
                    new_conn_str = swap_connected_printer_string(connected, pr_no, substitute_pr_no)
                    cursor.execute("UPDATE pcs SET connected_printers = ? WHERE id = ?", (new_conn_str, pc_id))

        # Update service record in database
        update_sql = """
            UPDATE printer_service SET 
                pr_no = ?, sla_no = ?, serial_no = ?, mac = ?, 
                model = ?, fault_description = ?, status = ?, 
                acquisition_date = ?, sent_date = ?, return_date = ?,
                has_substitute = ?, substitute_pr_no = ?, user_name = ?,
                location_code = ?
            WHERE id = ?
        """
        
        values = (
            pr_no, data.get('sla_no'), data.get('seri') or data.get('serial_no'), 
            data.get('mac'), data.get('model'), 
            fault_desc, data.get('status'), 
            acq_date, sent_date, return_date,
            has_substitute,
            substitute_pr_no,
            data.get('user_name', 'system'),
            current_location,
            record_id
        )
        cursor.execute(update_sql, values)

        # Update Printer status and location
        clean_loc = current_location.replace('SERVİSTE-', '').replace('DEPO-', '').replace('ARIZALI-', '').replace('KONTROLDE-', '').strip() or 'BİLİNMİYOR'
        final_new_location = current_location

        if is_returning:
            # Servisten dönen yazıcı kontrol deposuna (KONTROLDE) gider
            if has_substitute:
                kontrol_location = "KONTROLDE"
            else:
                kontrol_location = f"KONTROLDE-{clean_loc}"
                
            cursor.execute("UPDATE printers SET is_faulty=0, in_service=0, warehouse=0, on_field=0, location_code=? WHERE pr_no=?", (kontrol_location, pr_no))
            final_new_location = kontrol_location
            
            # Sütun eklendiğinde değeri sıfırla
            try:
                cursor.execute("UPDATE printer_service SET is_controlled=0 WHERE id=?", (record_id,))
            except Exception:
                pass
        else:
            status_str = str(data.get('status', '')).strip().upper()
            if status_str == 'SERVİSTE' or status_str == 'SERVISTE':
                if has_substitute:
                    new_location = "SERVİSTE"
                else:
                    new_location = f"SERVİSTE-{clean_loc}"
            elif status_str == 'ARIZALI' or status_str == 'ARIZALI':
                if has_substitute:
                    new_location = "ARIZALI"
                else:
                    new_location = f"ARIZALI-{clean_loc}"
            else:
                # Durum Tamamlandı. Zaten önceden dönmüş bir kayıt (is_returning=False).
                # is_controlled durumunu oku ve ona göre lokasyon belirle
                cursor.execute("SELECT is_controlled FROM printer_service WHERE id=?", (record_id,))
                ctrl_row = cursor.fetchone()
                is_ctrl = 0
                if ctrl_row and ctrl_row[0]:
                    is_ctrl = 1
                    
                if is_ctrl == 1:
                    if has_substitute:
                        new_location = "DEPO"
                    else:
                        new_location = f"DEPO-{clean_loc}"
                else:
                    if has_substitute:
                        new_location = "KONTROLDE"
                    else:
                        new_location = f"KONTROLDE-{clean_loc}"
                        
                cursor.execute("UPDATE printers SET is_faulty=0, in_service=0, warehouse=?, on_field=0, location_code=? WHERE pr_no=?", (1 if is_ctrl else 0, new_location, pr_no))
                
            final_new_location = new_location

            if has_substitute and substitute_pr_no:
                sub_row = find_printer_row(cursor, substitute_pr_no)
                if sub_row:
                    substitute_pr_no = sub_row[1]
                cursor.execute("UPDATE printers SET is_faulty=0, in_service=0, warehouse=0, on_field=1, location_code=? WHERE pr_no=?", (current_location, substitute_pr_no))
                cursor.execute("SELECT id FROM printers WHERE (pr_no=? OR pr_no=?) AND is_deleted=0", (substitute_pr_no, "".join(filter(str.isdigit, str(substitute_pr_no)))))
                sub_row = cursor.fetchone()
                if sub_row:
                    cursor.execute("UPDATE printers SET is_faulty=0, in_service=0, warehouse=0, on_field=1, location_code=? WHERE id=?", 
                                   (current_location, sub_row[0]))

        conn.commit()
        
        # Trigger CUPS synchronization
        try:
            from modules.inventory_printers import sync_printer_to_cups_internal
            sync_printer_to_cups_internal(pr_no, final_new_location, cursor, conn)
        except Exception as e:
            print(f"[CUPS SYNC ERROR during update_service] {e}")

        conn.close()

        # Trigger background automations
        user_id = request.current_user.get('user_id')
        
        final_original_loc = current_location
        if is_returning and substitute_pr_no and 'sub_location' in locals() and sub_location:
            final_original_loc = sub_location

        if is_returning:
            run_service_automations_async(
                user_id=user_id,
                pr_no=pr_no,
                substitute_pr_no=substitute_pr_no if (has_substitute or is_returning) else None,
                orig_location=final_original_loc,
                pc_targets=pc_targets,
                has_substitute=has_substitute,
                status_str=data['status'],
                is_return=is_returning
            )

        return jsonify({"success": True})
    except Exception as e:
        print("UPDATE SERVICE ERROR:", str(e))
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@service_manager_bp.route('/toggle_control/<int:record_id>', methods=['POST'])
@require_auth
def toggle_control(record_id):
    try:
        user_role = request.current_user.get('role', '')
        if user_role not in ('ADMIN', 'DEPOT'):
            return jsonify({"success": False, "error": "Bu işlem için yetkiniz bulunmamaktadır."}), 403

        data = request.json
        new_state = 1 if data.get('is_controlled') else 0

        # Eğer Depocu ise ve tiki kaldırmaya (0 yapmaya) çalışıyorsa engelle
        if user_role == 'DEPOT' and new_state == 0:
            return jsonify({"success": False, "error": "İşaretlenmiş bir kontrolü sadece Admin kaldırabilir!"}), 403

        conn = get_db_connection()
        if not conn:
            return jsonify({"success": False, "error": "Database connection failed"}), 500
        cursor = conn.cursor()

        # Tabloda is_controlled kolonu yoksa eklemeyi dene (İlk çalışma için)
        try:
            cursor.execute("ALTER TABLE printer_service ADD is_controlled BIT DEFAULT 0")
            conn.commit()
        except Exception:
            pass

        # Mevcut kaydı bul
        cursor.execute("SELECT pr_no, location_code, has_substitute FROM printer_service WHERE id = ?", (record_id,))
        row = cursor.fetchone()
        if not row:
            conn.close()
            return jsonify({"success": False, "error": "Kayıt bulunamadı."}), 404
            
        pr_no = row[0]
        existing_loc = row[1] or 'BİLİNMİYOR'
        has_substitute = bool(row[2])
        clean_loc = existing_loc.replace('SERVİSTE-', '').replace('DEPO-', '').replace('ARIZALI-', '').replace('KONTROLDE-', '').strip() or 'BİLİNMİYOR'

        # printer_service tablosunu güncelle
        cursor.execute("UPDATE printer_service SET is_controlled = ? WHERE id = ?", (new_state, record_id))

        final_new_location = existing_loc
        if new_state == 1:
            # Kontrol edildi (Tik atıldı) -> DEPO'ya al
            if has_substitute:
                depo_location = "DEPO"
            else:
                depo_location = f"DEPO-{clean_loc}"
            cursor.execute("UPDATE printers SET is_faulty=0, in_service=0, warehouse=1, on_field=0, location_code=? WHERE pr_no=?", (depo_location, pr_no))
            final_new_location = depo_location
        else:
            # Kontrol iptal edildi (Admin tiki kaldırdı) -> KONTROLDE'ye geri al
            if has_substitute:
                kontrol_location = "KONTROLDE"
            else:
                kontrol_location = f"KONTROLDE-{clean_loc}"
            cursor.execute("UPDATE printers SET is_faulty=0, in_service=0, warehouse=0, on_field=0, location_code=? WHERE pr_no=?", (kontrol_location, pr_no))
            final_new_location = kontrol_location

        conn.commit()
        
        # Trigger CUPS synchronization
        try:
            from modules.inventory_printers import sync_printer_to_cups_internal
            sync_printer_to_cups_internal(pr_no, final_new_location, cursor, conn)
        except Exception as e:
            print(f"[CUPS SYNC ERROR during toggle_control] {e}")

        conn.close()

        return jsonify({"success": True, "new_state": new_state})
    except Exception as e:
        print("TOGGLE CONTROL ERROR:", str(e))
        return jsonify({"success": False, "error": str(e)}), 500

@service_manager_bp.route('/delete/<int:record_id>', methods=['DELETE'])
@require_admin
def delete_service(record_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE printer_service SET is_deleted = 1 WHERE id = ?", (record_id,))
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@service_manager_bp.route('/export_pdf', methods=['GET'])
def export_pdf():
    import os
    import io
    from fpdf import FPDF
    from flask import send_file
    import datetime
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        query = """
            SELECT s.pr_no, s.serial_no, s.mac, s.model, s.fault_description, s.status, s.sent_date, s.return_date
            FROM printer_service s
            WHERE s.is_deleted = 0 
              AND (s.return_date IS NULL OR CAST(s.return_date AS DATE) = CAST(GETDATE() AS DATE))
            ORDER BY CASE WHEN s.acquisition_date IS NULL THEN 0 ELSE 1 END DESC, s.acquisition_date DESC
        """
        cursor.execute(query)
        rows = cursor.fetchall()
        
        # Fetch printer fallback data to fill empty serials/macs and add printers missing service records
        cursor.execute("SELECT pr_no, model, serial_no, mac, location_code, is_faulty, in_service FROM printers WHERE is_deleted = 0")
        printers_db = cursor.fetchall()
        conn.close()

        printers_map = {}
        for p_row in printers_db:
            p_pr_no, p_model, p_serial, p_mac, p_loc, p_is_faulty, p_in_service = p_row
            num = extract_printer_number(p_pr_no)
            if num is not None:
                printers_map[num] = {
                    'pr_no': p_pr_no,
                    'model': p_model,
                    'serial_no': p_serial,
                    'mac': p_mac,
                    'location_code': p_loc,
                    'is_faulty': p_is_faulty,
                    'in_service': p_in_service
                }

        # Convert to list of lists for mutability
        rows_list = [list(r) for r in rows]
        service_pr_nums = set()
        today_str = datetime.datetime.now().strftime("%Y-%m-%d")
        
        # 1. Fill empty fields in existing service records
        for r in rows_list:
            pr_no = r[0]
            num = extract_printer_number(pr_no)
            if num is not None:
                service_pr_nums.add(num)
                p_detail = printers_map.get(num)
                if p_detail:
                    if not r[1] or str(r[1]).strip() in ('', 'None', 'null', '-', '0', '0.0'):
                        r[1] = p_detail['serial_no'] or ''
                    if not r[2] or str(r[2]).strip() in ('', 'None', 'null', '-', '0', '0.0'):
                        r[2] = p_detail['mac'] or ''
                    if not r[3] or str(r[3]).strip() in ('', 'None', 'null', '-', '0', '0.0'):
                        r[3] = p_detail['model'] or ''

            # Status updates for today's transactions
            sent_d = r[6]
            ret_d = r[7]
            if sent_d and str(sent_d).startswith(today_str):
                r[5] = "Servise Teslim Edildi"
            if ret_d and str(ret_d).startswith(today_str):
                r[5] = "Teslim Alındı"

        # 2. Add printers marked as faulty/in_service that do not have active service records
        for num, p_detail in printers_map.items():
            if num not in service_pr_nums:
                is_faulty = p_detail['is_faulty'] == 1 or p_detail['is_faulty'] == True
                in_service = p_detail['in_service'] == 1 or p_detail['in_service'] == True
                if is_faulty or in_service:
                    status_str = 'Arızalı' if is_faulty else 'Serviste'
                    pr_display = p_detail['pr_no']
                    if str(pr_display).isdigit():
                        pr_display = f"PR-{str(pr_display).zfill(3)}"
                    
                    rows_list.append([
                        pr_display,
                        p_detail['serial_no'] or '',
                        p_detail['mac'] or '',
                        p_detail['model'] or '',
                        'Servis Kaydı Bulunmamaktadır',
                        status_str,
                        None,
                        None
                    ])
        
        rows = rows_list

        class PDF(FPDF):
            def header(self):
                # Font ayari
                if os.path.exists("C:\\Windows\\Fonts\\arial.ttf"):
                    self.add_font('Arial', '', 'C:\\Windows\\Fonts\\arial.ttf')
                    self.add_font('Arial', 'B', 'C:\\Windows\\Fonts\\arialbd.ttf')
                    
                # Ust border cizgisi (Mavi)
                self.set_line_width(0.5)
                self.set_draw_color(30, 64, 175)
                self.line(10, 10, 287, 10)
                
                # Sol Alan (KEYDATA LOGO YAZISI)
                self.set_y(15)
                self.set_font('Arial' if os.path.exists("C:\\Windows\\Fonts\\arialbd.ttf") else 'helvetica', 'B', 22)
                self.set_text_color(55, 71, 79)
                self.cell(40, 10, 'KEY', border=0, align='R')
                self.set_text_color(211, 47, 47)
                self.cell(20, 10, 'DATA', border=0, align='L')
                
                # Orta Alan (BASLIK)
                self.set_text_color(0, 0, 0)
                self.set_x(100)
                self.set_font('Arial' if os.path.exists("C:\\Windows\\Fonts\\arialbd.ttf") else 'helvetica', 'B', 18)
                self.cell(100, 10, 'YAZICI SERVIS TESLIM FORMU', border=0, align='C')
                
                # Sag Alan (Tarih)
                self.set_font('Arial' if os.path.exists("C:\\Windows\\Fonts\\arialbd.ttf") else 'helvetica', 'B', 14)
                self.set_x(247)
                self.cell(40, 10, f'{datetime.datetime.now().strftime("%d.%m.%Y")}', border=0, align='R')
                
                self.ln(12)
                
                # Alt border cizgisi (Mavi)
                self.set_draw_color(30, 64, 175)
                self.line(10, self.get_y(), 287, self.get_y())
                self.ln(5)

            def footer(self):
                # Imza alanlari
                self.set_y(-42)
                self.set_draw_color(180, 180, 180)
                self.line(10, self.get_y(), 287, self.get_y())
                self.ln(3)
                self.set_text_color(0, 0, 0)
                font_to_use = 'Arial' if os.path.exists("C:\\Windows\\Fonts\\arial.ttf") else 'helvetica'
                self.set_font(font_to_use, 'B', 10)
                
                self.set_x(50)
                self.cell(95, 10, 'Teslim Eden', border=0, align='C')
                self.set_x(150)
                self.cell(95, 10, 'Teslim Alan', border=0, align='C')
                
                self.set_y(-25)
                self.set_font(font_to_use, '', 9)
                self.set_x(50)
                self.cell(95, 10, 'Ad Soyad / Imza', border=0, align='C')
                self.set_x(150)
                self.cell(95, 10, 'Ad Soyad / Imza', border=0, align='C')
                
                # Sayfa numarasi
                self.set_y(-12)
                self.set_font(font_to_use, '', 8)
                self.set_text_color(100, 100, 100)
                self.cell(0, 10, f'Sayfa {self.page_no()}', 0, 0, 'C')

        pdf = PDF(orientation='L', format='A4')
        pdf.set_auto_page_break(True, margin=45)
        pdf.add_page()
        
        has_arial = os.path.exists("C:\\Windows\\Fonts\\arial.ttf")
        font_name = 'Arial' if has_arial else 'helvetica'
        
        # SUTUN GENISLIKLERI (Toplam 277mm - A4 Yatay Kullanilabilir Alan)
        col_widths = [10, 25, 35, 35, 40, 95, 37]
        headers = ['NO', 'PR NO', 'SERI NO', 'MAC ADRESI', 'YAZICI MODELI', 'ARIZA ACIKLAMASI', 'TESLIMAT DURUMU']
        
        pdf.set_font(font_name, 'B', 10)
        pdf.set_fill_color(240, 240, 240)
        pdf.set_text_color(0, 0, 0)
        pdf.set_draw_color(150, 150, 150)
        pdf.set_line_width(0.2)
        
        def add_table_header():
            pdf.set_font(font_name, 'B', 10)
            pdf.set_fill_color(240, 240, 240)
            pdf.set_text_color(0, 0, 0)
            for i, header in enumerate(headers):
                pdf.cell(col_widths[i], 12, header, border=1, align='C', fill=True)
            pdf.ln()
            pdf.set_font(font_name, '', 9)

        # Tablo Basliklari
        add_table_header()

        pdf.set_font(font_name, '', 9)
        
        # Tablo Icerigi
        for idx, row in enumerate(rows):
            no_str = str(idx + 1)
            pr_no = str(row[0] or '')
            seri = str(row[1] or '')
            mac = str(row[2] or '')
            model = str(row[3] or '')[:30]
            desc = str(row[4] or '') # Truncation kaldirildi
            
            # Dinamik Status Hesaplama
            status_raw = str(row[5] or '')
            if not status_raw or status_raw == 'None':
                if row[7]: # return_date
                    status_raw = 'Tamamlandı'
                elif row[6]: # sent_date
                    status_raw = 'Serviste'
                else:
                    status_raw = 'Arızalı'
            
            # Durum metni ve rengi
            if status_raw == 'Arızalı' or status_raw == 'Arizali':
                status_text = 'SERVISE TESLIM BEKLIYOR'
                status_color = (41, 128, 185) # Mavi ton
            elif status_raw == 'Serviste':
                status_text = 'SERVISTE'
                status_color = (21, 101, 192) # Koyu Mavi
            else:
                status_text = status_raw.upper()
                status_color = (100, 100, 100)

            # Satir Yuksekligi (Tek satir icin)
            min_row_h = 8
            text_h = 4  # Coklu satirlar arasi yazi yuksekligi
            
            # Aciklama icin satir sayisini hesapla
            pdf.set_font(font_name, '', 9)
            lines = 0
            for paragraph in desc.split('\n'):
                words = paragraph.split()
                if not words:
                    lines += 1
                    continue
                current_line = ""
                p_lines = 1
                for word in words:
                    if pdf.get_string_width(current_line + word) > col_widths[5] - 2:
                        p_lines += 1
                        current_line = word + " "
                    else:
                        current_line += word + " "
                lines += p_lines
            
            if lines == 0:
                lines = 1
            
            row_h = max(min_row_h, (text_h * lines) + 2)

            # Sayfa sonu kontrolu: tablo imza footer alanina girmesin.
            if pdf.get_y() + row_h > pdf.page_break_trigger:
                pdf.add_page()
                add_table_header()
            
            pdf.set_text_color(0, 0, 0)
            
            # 0. NO (Beyaz)
            pdf.set_fill_color(255, 255, 255)
            pdf.cell(col_widths[0], row_h, no_str, border=1, align='C', fill=True)
            
            # 1. PR NO (Acik Yesil Arkaplan)
            pdf.set_fill_color(200, 230, 201) 
            pdf.cell(col_widths[1], row_h, pr_no, border=1, align='C', fill=True)
            
            # Beyaza don
            pdf.set_fill_color(255, 255, 255)
            
            # 2. SERI NO
            pdf.cell(col_widths[2], row_h, seri, border=1, align='C', fill=True)
            
            # 3. MAC ADRESI
            pdf.cell(col_widths[3], row_h, mac, border=1, align='C', fill=True)
            
            # 4. YAZICI MODELI
            pdf.cell(col_widths[4], row_h, model, border=1, align='C', fill=True)
            
            # 5. ARIZA ACIKLAMASI (MultiCell ile kaydirma)
            x = pdf.get_x()
            y = pdf.get_y()
            # Arka plan ve cerceveyi ciz
            pdf.cell(col_widths[5], row_h, "", border=1, fill=True)
            # Metni icine yazdir (multi_cell y eksenini asagi kaydirir)
            pdf.set_xy(x, y + (row_h - (text_h * lines)) / 2) # Dikey ortalama icin hafif pay birak
            pdf.multi_cell(col_widths[5], text_h, desc, border=0, align='L', fill=False)
            
            # X, Y koordinatlarini hucre sonuna geri al
            pdf.set_xy(x + col_widths[5], y)
            
            # 6. TESLIMAT DURUMU (Mavi yazi rengi)
            pdf.set_text_color(*status_color)
            pdf.set_font(font_name, 'B', 8)
            pdf.cell(col_widths[6], row_h, status_text, border=1, align='C', fill=True)
            
            # Satir sonu, ayarlari sifirla ve sonraki satira gec
            pdf.set_text_color(0, 0, 0)
            pdf.set_font(font_name, '', 9)
            pdf.set_xy(10, y + row_h)
            
        pdf_bytes = pdf.output()
        
        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype='application/pdf',
            as_attachment=False,
            download_name=f'Servis_Teslim_Formu_{datetime.datetime.now().strftime("%Y%m%d")}.pdf'
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@service_manager_bp.route('/export_form', methods=['GET'])
@require_auth
def export_form():
    """Servis teslim formunu Excel (.xlsx) olarak döner."""
    import io
    import datetime
    from flask import send_file
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        return jsonify({"success": False, "error": "openpyxl kütüphanesi sunucuda yüklü değil."}), 500
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        query = """
            SELECT s.pr_no, s.serial_no, s.mac, s.model, s.fault_description, 
                   s.status, s.acquisition_date, s.sent_date, s.return_date,
                   s.location_code
            FROM printer_service s
            WHERE s.is_deleted = 0 
              AND (s.return_date IS NULL OR CAST(s.return_date AS DATE) = CAST(GETDATE() AS DATE))
            ORDER BY CASE WHEN s.acquisition_date IS NULL THEN 0 ELSE 1 END DESC, s.acquisition_date DESC
        """
        cursor.execute(query)
        columns = [col[0].lower() for col in cursor.description]
        rows = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        # Fetch printer fallback data to fill empty serials/macs and add printers missing service records
        cursor.execute("SELECT pr_no, model, serial_no, mac, location_code, is_faulty, in_service FROM printers WHERE is_deleted = 0")
        printers_db = cursor.fetchall()
        conn.close()

        printers_map = {}
        for p_row in printers_db:
            p_row_pr_no, p_model, p_serial, p_mac, p_loc, p_is_faulty, p_in_service = p_row
            num = extract_printer_number(p_row_pr_no)
            if num is not None:
                printers_map[num] = {
                    'pr_no': p_row_pr_no,
                    'model': p_model,
                    'serial_no': p_serial,
                    'mac': p_mac,
                    'location_code': p_loc,
                    'is_faulty': p_is_faulty,
                    'in_service': p_in_service
                }

        service_pr_nums = set()
        today_str = datetime.datetime.now().strftime("%Y-%m-%d")
        
        # 1. Fill empty fields in existing service records
        for r in rows:
            pr_no = r.get('pr_no')
            num = extract_printer_number(pr_no)
            if num is not None:
                service_pr_nums.add(num)
                p_detail = printers_map.get(num)
                if p_detail:
                    if not r.get('serial_no') or str(r['serial_no']).strip() in ('', 'None', 'null', '-', '0', '0.0'):
                        r['serial_no'] = p_detail['serial_no'] or ''
                    if not r.get('mac') or str(r['mac']).strip() in ('', 'None', 'null', '-', '0', '0.0'):
                        r['mac'] = p_detail['mac'] or ''
                    if not r.get('model') or str(r['model']).strip() in ('', 'None', 'null', '-', '0', '0.0'):
                        r['model'] = p_detail['model'] or ''
                    if not r.get('location_code') or str(r['location_code']).strip() in ('', 'None', 'null', '-', '0', '0.0'):
                        r['location_code'] = p_detail['location_code'] or ''

            # Status updates for today's transactions
            sent_d = r.get('sent_date')
            ret_d = r.get('return_date')
            if sent_d and str(sent_d).startswith(today_str):
                r['status'] = 'Servise Teslim Edildi'
            if ret_d and str(ret_d).startswith(today_str):
                r['status'] = 'Teslim Alındı'

        # 2. Add printers marked as faulty/in_service that do not have active service records
        for num, p_detail in printers_map.items():
            if num not in service_pr_nums:
                is_faulty = p_detail['is_faulty'] == 1 or p_detail['is_faulty'] == True
                in_service = p_detail['in_service'] == 1 or p_detail['in_service'] == True
                if is_faulty or in_service:
                    status_str = 'Arızalı' if is_faulty else 'Serviste'
                    pr_display = p_detail['pr_no']
                    if str(pr_display).isdigit():
                        pr_display = f"PR-{str(pr_display).zfill(3)}"
                    
                    rows.append({
                        'pr_no': pr_display,
                        'serial_no': p_detail['serial_no'] or '',
                        'mac': p_detail['mac'] or '',
                        'model': p_detail['model'] or '',
                        'fault_description': 'Servis Kaydı Bulunmamaktadır',
                        'status': status_str,
                        'acquisition_date': None,
                        'sent_date': None,
                        'return_date': None,
                        'location_code': p_detail['location_code'] or ''
                    })
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Servis Teslim Formu"
        
        # Başlık satırı
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = f"YAZICI SERVİS TESLİM FORMU — {datetime.datetime.now().strftime('%d.%m.%Y')}"
        title_cell.font = Font(name='Arial', size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Tablo başlıkları
        headers = ['NO', 'PR NO', 'SERİ NO', 'MAC ADRESİ', 'YAZICI MODELİ', 'ARIZA AÇIKLAMASI', 'DURUM', 'MAHAL']
        header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        ws.row_dimensions[3].height = 22
        
        # Kolon genişlikleri
        col_widths = [5, 10, 18, 20, 25, 45, 18, 20]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w
        
        # Veri satırları
        data_font = Font(name='Arial', size=9)
        data_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for row_idx, record in enumerate(rows, 4):
            # Durum hesapla
            status = record.get('status') or ''
            if not status:
                if record.get('return_date'):
                    status = 'Tamamlandı'
                elif record.get('sent_date'):
                    status = 'Serviste'
                else:
                    status = 'Arızalı'
            
            if status in ('Arızalı', 'Arizali'):
                status_text = 'SERVİSE TESLİM BEKLİYOR'
            elif status == 'Serviste':
                status_text = 'SERVİSTE'
            else:
                status_text = status.upper()
            
            values = [
                row_idx - 3,
                record.get('pr_no', ''),
                record.get('serial_no', ''),
                record.get('mac', ''),
                record.get('model', ''),
                record.get('fault_description', ''),
                status_text,
                record.get('location_code', '')
            ]
            
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val if val else '')
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
                
                # Durum sütunu renklendirme
                if col_idx == 7:
                    if 'BEKLİYOR' in str(val):
                        cell.font = Font(name='Arial', size=9, bold=True, color='2980B9')
                    elif 'SERVİSTE' in str(val):
                        cell.font = Font(name='Arial', size=9, bold=True, color='1565C0')
        
        # İmza alanları
        last_row = len(rows) + 6
        ws.merge_cells(f'B{last_row}:C{last_row}')
        ws.merge_cells(f'F{last_row}:G{last_row}')
        ws.cell(row=last_row, column=2, value='Teslim Eden').font = Font(name='Arial', size=11, bold=True)
        ws.cell(row=last_row, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=last_row, column=6, value='Teslim Alan').font = Font(name='Arial', size=11, bold=True)
        ws.cell(row=last_row, column=6).alignment = Alignment(horizontal='center')
        
        ws.merge_cells(f'B{last_row+2}:C{last_row+2}')
        ws.merge_cells(f'F{last_row+2}:G{last_row+2}')
        ws.cell(row=last_row+2, column=2, value='Ad Soyad / İmza').font = Font(name='Arial', size=10)
        ws.cell(row=last_row+2, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=last_row+2, column=6, value='Ad Soyad / İmza').font = Font(name='Arial', size=10)
        ws.cell(row=last_row+2, column=6).alignment = Alignment(horizontal='center')
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'Servis_Teslim_Formu_{datetime.datetime.now().strftime("%Y%m%d")}.xlsx'
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500
