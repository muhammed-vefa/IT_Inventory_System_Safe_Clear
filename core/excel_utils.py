import openpyxl
import os

def read_excel_data(file_path, headers=None, sheet_name=None):
    """Excel dosyasını okuyup liste döndürür."""
    if not os.path.exists(file_path):
        return []
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if sheet_name is not None:
            if isinstance(sheet_name, int):
                try:
                    sheet = wb.worksheets[sheet_name]
                except IndexError:
                    return []
            elif sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
            else:
                # Case-insensitive match
                for name in wb.sheetnames:
                    if str(name).lower() == str(sheet_name).lower():
                        sheet = wb[name]
                        break
                else:
                    return [] # Bulunamadıysa boş liste dön
        else:
            sheet = wb.active
            
        data = []
        
        # Başlıkları al (eğer verilmemişse ilk satırı kullan)
        rows = list(sheet.rows)
        if not rows:
            return []
            
        if headers is None:
            headers = [str(cell.value).strip() if cell.value else f"Col{i}" for i, cell in enumerate(rows[0])]
            start_row = 1
        else:
            start_row = 0
            
        for row in rows[start_row:]:
            item = {}
            for i, cell in enumerate(row):
                if i < len(headers):
                    item[headers[i]] = cell.value
            if any(item.values()): # Boş satırları atla
                data.append(item)
        return data
    except Exception as e:
        print(f"Excel Okuma Hatası ({file_path}): {e}")
        return []

def write_excel_data(file_path, data, headers):
    """Veriyi Excel dosyasına geri yazar (Senkronizasyon)."""
    try:
        wb = openpyxl.Workbook()
        sheet = wb.active
        
        # Başlıkları yaz
        for i, h in enumerate(headers):
            sheet.cell(row=1, column=i+1, value=h)
            
        # Veriyi yaz
        for r_idx, item in enumerate(data):
            for c_idx, h in enumerate(headers):
                sheet.cell(row=r_idx+2, column=c_idx+1, value=item.get(h))
                
        wb.save(file_path)
        return True
    except Exception as e:
        print(f"Excel Yazma Hatası ({file_path}): {e}")
        return False
