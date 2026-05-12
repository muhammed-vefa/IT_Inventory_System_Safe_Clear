import openpyxl
import os

path = r'C:\Users\MUHAMMED-VEFA-IS\.gemini\antigravity\scratch\IT_Inventory_System\database\ana_database\yazıcılar.xlsx'

if not os.path.exists(path):
    print(f"File not found: {path}")
    exit()

wb = openpyxl.load_workbook(path, data_only=True)
print(f"Sheets: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    print(f"\n--- SHEET: {sheet_name} ---")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("Empty sheet")
        continue
    
    headers = rows[0]
    print(f"Headers: {headers}")
    for i, row in enumerate(rows[1:6], 1):
        print(f"Row {i}: {row}")
