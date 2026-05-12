import os
import re

def update_file(path, replacements):
    with open(path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    for old, new in replacements.items():
        content = content.replace(old, new)
        
    with open(path, 'w', encoding='utf-8') as f:
        f.write(content)

# 1. main.py Güncellemesi
main_path = r'C:\Users\MUHAMMED-VEFA-IS\.gemini\antigravity\scratch\IT_Inventory_System\main.py'
main_new_sync = """def sync_excel_to_db_internal():
    \"\"\"Excel verilerini Bilgi Bankası'nın (notes_manager) sağlam mantığıyla senkronize eder.\"\"\"
    from core.database_sql import get_db_connection, init_db
    from core.excel_utils import read_excel_data
    from core.utils import _clean, _norm_key, _get
    import logging

    logging.info("DEBUG: Kapsamlı Excel Senkronizasyonu (v8.5) başlatılıyor...")
    stats = {"pc_synced": 0, "areas_synced": 0, "keyos_errors": 0}

    # Dosya yollarını kesinleştir
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    def get_p(filename): return os.path.join(BASE_DIR, 'database', 'ana_database', filename)

    mt_path    = get_p("mahal_telefon.xlsx")
    env_path   = get_p("envanter.xlsx")
    alan_path  = get_p("ORTAK_ALANLAR.xlsx")

    init_db()
    conn = get_db_connection()

    # 1. Mahal Cache (Bilgi Bankası stili esnek okuma)
    mahal_cache = {}
    if os.path.exists(mt_path):
        data = read_excel_data(mt_path, sheet_name=0)
        for m in data:
            mk = _clean(_get(m, ['MAHAL', 'MAHAL KODU', 'KOD']))
            if mk:
                mahal_cache[mk] = {
                    'adi': _clean(_get(m, ['MAHAL ADI', 'ADI', 'LOKASYON'])),
                    'tel': _clean(_get(m, ['DAHİLİ', 'TELEFON', 'TEL']))
                }

    # 2. Envanter (PC, Tablet, Sıramatik) - Esnek Sayfa ve Sütun Mantığı
    if os.path.exists(env_path):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(env_path, data_only=True)
            conn.execute("DELETE FROM inventory")
            
            for sheet_name in wb.sheetnames:
                s_up = _norm_key(sheet_name)
                if not any(x in s_up for x in ['BILGISAYAR', 'TABLET', 'SIRAMATIK', 'KIOSK', 'PC', 'ENVANTER']): continue
                
                data = read_excel_data(env_path, sheet_name=sheet_name)
                d_type = 'PC'
                if 'TABLET' in s_up: d_type = 'TABLET'
                elif 'SIRAMATIK' in s_up or 'KIOSK' in s_up: d_type = 'SIRAMATIK'

                for item in data:
                    pc_no = _clean(_get(item, ['PC', 'PC NO', 'ID', 'PC NUMARASI']))
                    if not pc_no: continue
                    
                    mk = _clean(_get(item, ['MAHAL KODU', 'MAHAL', 'KOD']))
                    m = mahal_cache.get(mk, {'adi': '', 'tel': ''})
                    
                    keyos = 1 if str(_get(item, ['KEYOS', 'KOS', 'K-OS']) or '').upper() in ['1', 'TRUE', 'VAR', 'EVET', 'X', '*'] else 0
                    windows = 1 if not keyos and str(_get(item, ['WINDOWS', 'WIN']) or '').upper() in ['1', 'TRUE', 'VAR', 'EVET', 'X', '*'] else 0
                    
                    conn.execute('''INSERT INTO inventory (
                        pc_no, mahal_kodu, mahal_adi, ip, mac, bagli_yazicilar, 
                        windows, keyos, device_type, card_name, assigned_to, phone, title, unit, aciklama
                    ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', (
                        pc_no, mk, m['adi'], 
                        _clean(_get(item, ['İP', 'IP', 'IP ADRESİ'])),
                        _clean(_get(item, ['MAC ADRES', 'MAC', 'MAC ADRESİ'])),
                        str(_get(item, ['BAĞLI OLAN YAZICILAR', 'YAZICILAR', 'YAZICI']) or ''),
                        windows, keyos, d_type,
                        _clean(_get(item, ['KART ADI', 'CİHAZ ADI', 'ADI', 'ENVANTER KART ADI'])),
                        _clean(_get(item, ['ZİMMETLENEN KİŞİ', 'ZIMMET', 'PERSONEL', 'KULLANICI'])),
                        _clean(_get(item, ['CEP TELEFON', 'TELEFON'])) or m['tel'],
                        _clean(_get(item, ['UNVAN', 'GÖREV', 'ÜNVAN'])),
                        _clean(_get(item, ['BİRİM', 'BIRIM', 'UNIT'])),
                        str(_get(item, ['AÇIKLAMA', 'NOT']) or '')
                    ))
                    stats["pc_synced"] += 1
            conn.commit()
        except Exception as e:
            logging.error(f"Envanter Sync Hatası: {e}")

    # 3. Ortak Alanlar
    if os.path.exists(alan_path):
        try:
            conn.execute("DELETE FROM shared_areas")
            data = read_excel_data(alan_path, sheet_name=0) # İlk sayfa
            for r in data:
                name = _get(r, ['NAME', 'AD', 'ALAN ADI'])
                if name:
                    conn.execute("INSERT INTO shared_areas (name, [user], password, path) VALUES (?,?,?,?)", (
                        name, _get(r, ['USER', 'KULLANICI']), _get(r, ['PASSWORD', 'ŞİFRE']), _get(r, ['PATH', 'YOL'])
                    ))
                    stats["areas_synced"] += 1
            conn.commit()
        except Exception as e:
            logging.error(f"Ortak Alanlar Sync Hatası: {e}")

    conn.close()
    return stats"""

with open(main_path, 'r', encoding='utf-8') as f:
    main_content = f.read()

pattern = re.compile(r'def sync_excel_to_db_internal\(\):.*?return stats', re.DOTALL)
main_content = pattern.sub(main_new_sync, main_content)

main_sync_all_new = """@app.route('/api/sync/all', methods=['POST'])
@require_admin
def sync_all():
    \"\"\"Tüm Excel dosyalarını senkronize eder.\"\"\"
    results = []
    success = True
    try:
        from modules.printer_manager import sync_printers_from_excel_internal
        from modules.depot_manager import sync_depot_from_excel_internal
        from modules.notes_manager import sync_kb_from_excel_internal
        
        # 1. Envanter & Ortak Alanlar
        try:
            inv_stats = sync_excel_to_db_internal()
            results.append(f"Envanter: {inv_stats.get('pc_synced', 0)} cihaz güncellendi.")
            results.append(f"Ortak Alanlar: {inv_stats.get('areas_synced', 0)} kayıt güncellendi.")
        except Exception as e:
            results.append(f"Envanter Hatası: {str(e)}")
            success = False

        # 2. Yazıcılar
        try:
            p_count = sync_printers_from_excel_internal()
            results.append(f"Yazıcılar: {p_count} kayıt senkronize edildi.")
        except Exception as e:
            results.append(f"Yazıcı Hatası: {str(e)}")
            success = False

        # 3. Diğerleri
        try:
            results.append(f"Depo: {sync_depot_from_excel_internal()} ürün güncellendi.")
            results.append(f"Bilgi Bankası: {sync_kb_from_excel_internal()} kayıt güncellendi.")
        except: success = False

        # Yanıtı UTF-8 olarak gönder
        resp = jsonify({
            "success": success,
            "message": "Toplu senkronizasyon tamamlandı." if success else "Bazı modüllerde hatalar oluştu.",
            "details": results
        })
        resp.charset = 'utf-8'
        return resp
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500"""

pattern_sync_all = re.compile(r'@app\.route\(\'/api/sync/all\'.*?return jsonify\(.*?500', re.DOTALL)
main_content = pattern_sync_all.sub(main_sync_all_new, main_content)

with open(main_path, 'w', encoding='utf-8') as f:
    f.write(main_content)

pm_path = r'C:\Users\MUHAMMED-VEFA-IS\.gemini\antigravity\scratch\IT_Inventory_System\modules\printer_manager.py'
pm_new_sync = """def sync_printers_from_excel_internal():
    \"\"\"Excel'den tüm yazıcı, tarayıcı, barkod ve servis verilerini Bilgi Bankası'nın sağlam mantığıyla senkronize eder.\"\"\"
    from core.database_sql import get_db_connection
    from core.excel_utils import read_excel_data
    from core.utils import _clean, _norm_key, _get
    import logging
    import openpyxl

    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.abspath(os.path.join(BASE_DIR, '..', 'database', 'ana_database', 'yazıcılar.xlsx'))
    
    if not os.path.exists(excel_path): return 0

    try:
        conn = get_db_connection()
        conn.execute("DELETE FROM printer_service")
        conn.execute("DELETE FROM printers")
        conn.commit()

        wb = openpyxl.load_workbook(excel_path, data_only=True)
        p_count = 0
        s_count = 0

        for sheet_name in wb.sheetnames:
            data = read_excel_data(excel_path, sheet_name=sheet_name)
            if not data: continue
            s_up = _norm_key(sheet_name)
            is_service = any(x in s_up for x in ['SERVIS', 'HISTORY', 'ARIZA', 'ISLEM', 'IŞLEM'])
            is_device = any(x in s_up for x in ['YAZICI', 'BARKOD', 'TARAYICI', 'OKUYUCU'])

            for item in data:
                if is_service:
                    pr_no = _clean(_get(item, ['Yazıcı Numarası', 'PR NO', 'YAZICI NO', 'ID', 'NO', 'YAZICI NUMARASI', 'PR NUMARASI']))
                    if not pr_no: continue
                    acq_date = str(_get(item, ['Alındığı Tarih', 'ALINDI TARIH', 'ALINAN TARİH', 'ACQ DATE']) or '')
                    sent_date = str(_get(item, ['Servise Gittiği Tarih', 'SERVISE GITTIGI TARIH', 'GİTTİĞİ TARİH', 'DATE', 'TARIH']) or '')
                    ret_date = str(_get(item, ['Servisten Geldiği Tarih', 'SERVISTEN GELDIGI TARIH', 'GELDİĞİ TARİH', 'RETURN']) or '')
                    srv_status = "Arızalı"
                    if ret_date and ret_date.strip(): srv_status = "Tamamlandı"
                    elif sent_date and sent_date.strip(): srv_status = "Serviste"
                    elif not acq_date.strip(): srv_status = _get(item, ['DURUM', 'STATUS', 'DURUMU']) or "Arızalı"

                    conn.execute("INSERT INTO printer_service (pr_no, seri, mac, fault_desc, status, sent_date, acq_date, return_date, mahal, substitute_pr_no) VALUES (?,?,?,?,?,?,?,?,?,?)",
                        (pr_no, _clean(_get(item, ['Seri Numarası', 'SERI NUMARASI', 'SERİ NUMARASI', 'SERIAL', 'SERI', 'SERİ'])), 
                         _clean(_get(item, ['MAC ADRESİ', 'MAC ADRESI', 'MAC', 'MAC ADRESS'])),
                         str(_get(item, ['Arıza Açıklaması', 'ARIZA ACIKLAMASI', 'ARIZA AÇIKLAMASI', 'HATA', 'ARIZA']) or ''),
                         srv_status, sent_date, acq_date, ret_date,
                         _clean(_get(item, ['Mahal', 'MAHAL', 'LOKASYON'])),
                         _clean(_get(item, ['İkame Yazıcı', 'IKAME YAZICI', 'İKAME YAZICI'])))
                    )
                    if srv_status == "Tamamlandı":
                        conn.execute("UPDATE printers SET status = 'Depo' WHERE pr_no = ?", (pr_no,))
                    s_count += 1
                elif is_device:
                    pr_no = _clean(_get(item, ['KAYITLI CİHAZ NO', 'KAYITLI CIHAZ NO', 'PR NUMARASI', 'PR NO', 'ID', 'NO']))
                    seri = _clean(_get(item, ['SERİ NUMARASI', 'SERI NUMARASI', 'SERIAL', 'SERI', 'SERİ']))
                    if not pr_no and not seri: continue
                    model = _clean(_get(item, ['CİHAZ ADI', 'CIHAZ ADI', 'ENVANTER KART ADI', 'MODEL', 'TUR', 'TÜR'])) or sheet_name
                    conn.execute("INSERT INTO printers (pr_no, seri, model, mac, ip, status, mahal) VALUES (?,?,?,?,?,?,?)", (
                        pr_no, seri, model,
                        _clean(_get(item, ['MAC ADRESS', 'MAC', 'MAC ADRESİ', 'MAC ADRESI'])),
                        _clean(_get(item, ['IP ADRES', 'IP', 'IP ADRESI', 'IP ADRESS'])), 
                        _clean(_get(item, ['DURUM', 'STATUS', 'DURUMU'])) or 'Kurulu',
                        _clean(_get(item, ['MAHAL', 'LOKASYON', 'BIRIM']))
                    ))
                    p_count += 1
        conn.commit()
        conn.close()
        return p_count + s_count
    except Exception as e:
        logging.error(f"printer_manager Sync Kritik Hata: {e}")
        return 0"""

with open(pm_path, 'r', encoding='utf-8') as f:
    pm_content = f.read()

pattern_pm = re.compile(r'def sync_printers_from_excel_internal\(\):.*?return 0', re.DOTALL)
pm_content = pattern_pm.sub(pm_new_sync, pm_content)

with open(pm_path, 'w', encoding='utf-8') as f:
    f.write(pm_content)

print("Guncelleme basariyla tamamlandi.")
