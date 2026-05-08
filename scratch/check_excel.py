import openpyxl
import os

excel_path = r'C:\Users\MUHAMMED-VEFA-IS\.gemini\antigravity\scratch\IT_Inventory_System\database\depo_envanteri.xlsx'
if os.path.exists(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    print("Sheets:", wb.sheetnames)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    print("Headers:", headers)
    
    # Kategori sütunundaki benzersiz değerleri bulalım (Eğer varsa)
    cat_idx = -1
    for i, h in enumerate(headers):
        if h and 'KATEGOR' in str(h).upper():
            cat_idx = i
            break
    
    if cat_idx != -1:
        cats = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[cat_idx]:
                cats.add(str(row[cat_idx]))
        print("Unique Categories in column:", cats)
else:
    print("Excel file not found at", excel_path)
