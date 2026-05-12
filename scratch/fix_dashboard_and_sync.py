import os
import re

main_path = r'C:\Users\MUHAMMED-VEFA-IS\.gemini\antigravity\scratch\IT_Inventory_System\main.py'

with open(main_path, 'r', encoding='utf-8') as f:
    content = f.read()

new_sync_logic = """def sync_excel_to_db_internal():
    \"\"\"Excel verilerini Bilgi Bankası'nın sağlam mantığı ve yedek dosyadaki akıllı eşleştirmeyle senkronize eder.\"\"\"
    from core.database_sql import get_db_connection, init_db
    from core.excel_utils import read_excel_data
    from core.utils import _clean, _norm_key, _get, _norm_pc_id
    import logging
    import openpyxl

    logging.info("DEBUG: Akıllı Senkronizasyon (v8.8) başlatılıyor...")
    stats = {"pc_synced": 0, "areas_synced": 0}

    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    def get_p(filename): return os.path.join(BASE_DIR, 'database', 'ana_database', filename)

    mt_path    = get_p("mahal_telefon.xlsx")
    env_path   = get_p("envanter.xlsx")
    alan_path  = get_p("ORTAK_ALANLAR.xlsx")
    pr_path    = get_p("yazıcılar.xlsx")

    init_db()
    conn = get_db_connection()

    mahal_cache = {}
    if os.path.exists(mt_path):
        data = read_excel_data(mt_path, sheet_name=0)
        for m in data:
            mk = _clean(_get(m, ['MAHAL', 'MAHAL KODU', 'KOD']))
            if mk:
                mahal_cache[mk] = {
                    'adi': _clean(_get(m, ['MAHAL ADI', 'ADI', 'LOKASYON'])),
                    'tel': _clean(_get(m, ['DAHİLİ', 'TELEFON', 'TEL']))
                }

    peripheral_cache = {}
    if os.path.exists(pr_path):
        try:
            wb_pr = openpyxl.load_workbook(pr_path, data_only=True)
            variants = [
                (['barkod_yazıcı', 'barkod yazıcı', 'by', 'Barkod Yazıcı'], 'by'),
                (['barkod_okuyucu', 'barkod okuyucu', 'bo', 'Barkod Okuyucu'], 'bo'),
                (['tarayıcı', 'tarayici', 'tr', 'Tarayıcı'], 'tr'),
            ]
            for sheets, key in variants:
                for sname in wb_pr.sheetnames:
                    s_norm = _norm_key(sname)
                    if any(_norm_key(x) in s_norm for x in sheets):
                        p_data = read_excel_data(pr_path, sheet_name=sname)
                        for p in p_data:
                            pc_id = _norm_pc_id(_get(p, ['KAYITLI CİHAZ NO', 'CİHAZ NO', 'PC NO', 'PC ID']))
                            seri = _clean(_get(p, ['SERİ NUMARASI', 'SERI NO', 'SERIAL']))
                            if pc_id and seri and seri not in ('0', 'NONE', '-'):
                                peripheral_cache.setdefault(pc_id, {})[key] = seri
        except: pass

    if os.path.exists(env_path):
        try:
            wb = openpyxl.load_workbook(env_path, data_only=True)
            conn.execute("DELETE FROM inventory")
            for sheet_name in wb.sheetnames:
                s_up = _norm_key(sheet_name)
                if not any(x in s_up for x in ['BILGISAYAR', 'TABLET', 'SIRAMATIK', 'KIOSK', 'PC', 'ENVANTER']): continue
                
                data = read_excel_data(env_path, sheet_name=sheet_name)
                d_type = 'PC'
                if 'TABLET' in s_up: d_type = 'TABLET'
                elif 'SIRAMATIK' in s_up or 'KIOSK' in s_up: d_type = 'SIRAMATIK'

                for item in data:
                    pc_no = _clean(_get(item, ['PC', 'PC NO', 'ID', 'PC NUMARASI']))
                    if not pc_no: continue
                    
                    norm_id = _norm_pc_id(pc_no)
                    p_info = peripheral_cache.get(norm_id, {})
                    
                    mk = _clean(_get(item, ['MAHAL KODU', 'MAHAL', 'KOD']))
                    m = mahal_cache.get(mk, {'adi': '', 'tel': ''})
                    
                    keyos = 1 if str(_get(item, ['KEYOS', 'KOS', 'K-OS']) or '').upper() in ['1', 'TRUE', 'VAR', 'EVET', 'X', '*'] else 0
                    windows = 1 if not keyos and str(_get(item, ['WINDOWS', 'WIN']) or '').upper() in ['1', 'TRUE', 'VAR', 'EVET', 'X', '*'] else 0
                    
                    by_seri = _clean(_get(item, ['BARKOD YAZICI SERİ NO', 'BY SERI'])) or p_info.get('by', '')
                    bo_seri = _clean(_get(item, ['BARKOD OKUYUCU SERİ NO', 'BO SERI'])) or p_info.get('bo', '')
                    tr_seri = _clean(_get(item, ['TARAYICI SERİ NO', 'TR SERI'])) or p_info.get('tr', '')

                    conn.execute('''INSERT INTO inventory (
                        pc_no, mahal_kodu, mahal_adi, ip, mac, bagli_yazicilar, 
                        windows, keyos, device_type, card_name, assigned_to, phone, title, unit, aciklama,
                        by_seri, bo_seri, tarayici_seri
                    ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', (
                        pc_no, mk, m['adi'], 
                        _clean(_get(item, ['İP', 'IP', 'IP ADRESİ'])),
                        _clean(_get(item, ['MAC ADRES', 'MAC', 'MAC ADRESİ'])),
                        str(_get(item, ['BAĞLI OLAN YAZICILAR', 'YAZICILAR', 'YAZICI']) or ''),
                        windows, keyos, d_type,
                        _clean(_get(item, ['KART ADI', 'CİHAZ ADI', 'ADI', 'ENVANTER KART ADI'])),
                        _clean(_get(item, ['ZİMMETLENEN KİŞİ', 'ZIMMET', 'PERSONEL', 'KULLANICI'])),
                        _clean(_get(item, ['CEP TELEFON', 'TELEFON'])) or m['tel'],
                        _clean(_get(item, ['UNVAN', 'GÖREV', 'ÜNVAN'])),
                        _clean(_get(item, ['BİRİM', 'BIRIM', 'UNIT'])),
                        str(_get(item, ['AÇIKLAMA', 'NOT']) or ''),
                        by_seri, bo_seri, tr_seri
                    ))
                    stats["pc_synced"] += 1
            conn.commit()
        except Exception as e: logging.error(f"Envanter Sync Hatası: {e}")

    if os.path.exists(alan_path):
        try:
            conn.execute("DELETE FROM shared_areas")
            data = read_excel_data(alan_path, sheet_name=0)
            for r in data:
                name = _get(r, ['NAME', 'AD', 'ALAN ADI'])
                if name:
                    conn.execute("INSERT INTO shared_areas (name, [user], password, path) VALUES (?,?,?,?)", (
                        name, _get(r, ['USER', 'KULLANICI']), _get(r, ['PASSWORD', 'ŞİFRE']), _get(r, ['PATH', 'YOL'])
                    ))
                    stats["areas_synced"] += 1
            conn.commit()
        except: pass

    conn.close()
    return stats"""

pattern = re.compile(r'def sync_excel_to_db_internal\(\):.*?return stats', re.DOTALL)
content = pattern.sub(new_sync_logic, content)

new_stats_logic = """@app.route('/api/dashboard/stats', methods=['GET'])
@require_auth
def get_dashboard_stats():
    conn = get_db_connection()
    pc_data = conn.execute(\"\"\"
        SELECT 
            COUNT(CASE WHEN windows=1 THEN 1 END) as win,
            COUNT(CASE WHEN keyos=1 THEN 1 END) as keyos,
            COUNT(CASE WHEN device_type='PC' THEN 1 END) as total_pc
        FROM inventory\"\"\").fetchone()

    pr_data = conn.execute(\"\"\"
        SELECT 
            COUNT(CASE WHEN status IN ('Kurulu', 'Sahada') THEN 1 END) as sahada,
            COUNT(CASE WHEN status='Depo' THEN 1 END) as depo,
            COUNT(CASE WHEN status='Arızalı' THEN 1 END) as ariza,
            COUNT(CASE WHEN status='Kayıp' THEN 1 END) as kayip
        FROM printers WHERE model NOT LIKE '%Barkod%' AND model NOT LIKE '%Tarayıcı%'\"\"\").fetchone()

    bo_kurulu = conn.execute(\"SELECT COUNT(*) FROM inventory WHERE bo_seri IS NOT NULL AND bo_seri != ''\").fetchone()[0]
    by_kurulu = conn.execute(\"SELECT COUNT(*) FROM inventory WHERE by_seri IS NOT NULL AND by_seri != ''\").fetchone()[0]
    
    bo_depo = conn.execute(\"SELECT COUNT(*) FROM printers WHERE model LIKE '%Barkod Okuyucu%' AND status='Depo'\").fetchone()[0]
    by_depo = conn.execute(\"SELECT COUNT(*) FROM printers WHERE model LIKE '%Barkod Yazıcı%' AND status='Depo'\").fetchone()[0]

    tr_c230_kurulu = conn.execute(\"SELECT COUNT(*) FROM inventory WHERE tarayici_seri LIKE '%C230%' OR tarayici_seri IN (SELECT seri FROM printers WHERE model LIKE '%C230%' AND status IN ('Kurulu', 'Sahada'))\").fetchone()[0]
    tr_g2090_kurulu = conn.execute(\"SELECT COUNT(*) FROM inventory WHERE tarayici_seri LIKE '%G2090%' OR tarayici_seri IN (SELECT seri FROM printers WHERE model LIKE '%G2090%' AND status IN ('Kurulu', 'Sahada'))\").fetchone()[0]
    
    tr_c230_depo = conn.execute(\"SELECT COUNT(*) FROM printers WHERE model LIKE '%C230%' AND status='Depo'\").fetchone()[0]
    tr_g2090_depo = conn.execute(\"SELECT COUNT(*) FROM printers WHERE model LIKE '%G2090%' AND status='Depo'\").fetchone()[0]

    return jsonify({
        'pc': {'win': pc_data['win'], 'keyos': pc_data['keyos']},
        'yazici': {'sahada': pr_data['sahada'], 'depo': pr_data['depo'], 'ariza': pr_data['ariza'], 'kayip': pr_data['kayip']},
        'bo': {'kurulu': bo_kurulu, 'depo': bo_depo},
        'by': {'kurulu': by_kurulu, 'depo': by_depo},
        'tr_c230': {'kurulu': tr_c230_kurulu, 'depo': tr_c230_depo},
        'tr_g2090': {'kurulu': tr_g2090_kurulu, 'depo': tr_g2090_depo}
    })"""

pattern_stats = re.compile(r'@app\.route\(\'/api/dashboard/stats\'.*?\}\)', re.DOTALL)
content = pattern_stats.sub(new_stats_logic, content)

with open(main_path, 'w', encoding='utf-8') as f:
    f.write(content)

pm_path = r'C:\Users\MUHAMMED-VEFA-IS\.gemini\antigravity\scratch\IT_Inventory_System\modules\printer_manager.py'
with open(pm_path, 'r', encoding='utf-8') as f:
    pm_content = f.read()

pm_new_logic = """                elif is_device:
                    pr_no = _clean(_get(item, ['KAYITLI CİHAZ NO', 'KAYITLI CIHAZ NO', 'PR NUMARASI', 'PR NO', 'ID', 'NO']))
                    seri = _clean(_get(item, ['SERİ NUMARASI', 'SERI NO', 'SERIAL']))
                    mac = _clean(_get(item, ['MAC ADRESS', 'MAC', 'MAC ADRESİ']))
                    
                    if not seri and not mac:
                        continue
                        
                    model = _clean(_get(item, ['CİHAZ ADI', 'CIHAZ ADI', 'ENVANTER KART ADI', 'MODEL', 'TUR', 'TÜR'])) or sheet_name"""

# Use a more robust replace for the printer_manager
pattern_pm = re.compile(r'elif is_device:.*?model = _clean\(_get\(item, \[\'CİHAZ ADI\'.*?\) or sheet_name', re.DOTALL)
pm_content = pattern_pm.sub(pm_new_logic, pm_content)

with open(pm_path, 'w', encoding='utf-8') as f:
    f.write(pm_content)

print("Dashboard and Sync updated successfully.")
