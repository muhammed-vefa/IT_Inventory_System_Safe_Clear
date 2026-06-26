import os
import sys
# Path patch for sub-folder execution
project_root = os.path.dirname(os.path.abspath(__file__))
if os.path.basename(project_root) == 'tools':
    project_root = os.path.dirname(project_root)
if project_root not in sys.path:
    sys.path.append(project_root)
os.chdir(project_root)

# -*- coding: utf-8 -*-
"""
IT_INVENTORY v12.9 — Excel'den SQL Server'a Zırhlı Veri Yükleme Motoru
"""
import os
import sys
import pyodbc
import openpyxl
import datetime
from dotenv import load_dotenv
from werkzeug.security import generate_password_hash

BASE_DIR = project_root
env_path = os.path.join(BASE_DIR, ".env")
if not os.path.exists(env_path):
    env_path = os.path.join(BASE_DIR, "tools", ".env")
load_dotenv(env_path, override=True)

DB_SERVER = os.getenv("DB_SERVER", ".\\SQLEXPRESS").strip()
DB_NAME = os.getenv("DB_NAME", "IT_INVENTORY").strip()
DB_USER = os.getenv("DB_USER", "").strip()
DB_PASS = os.getenv("DB_PASS", "").strip()

DATABASE_DIR = os.path.join(BASE_DIR, "database")

def get_connection():
    if DB_USER and DB_PASS:
        conn_str = f"DRIVER={{SQL Server}};SERVER={DB_SERVER};DATABASE={DB_NAME};UID={DB_USER};PWD={DB_PASS};"
    else:
        conn_str = f"DRIVER={{SQL Server}};SERVER={DB_SERVER};DATABASE={DB_NAME};Trusted_Connection=yes;"
    return pyodbc.connect(conn_str, timeout=10)

def get_table_columns(cursor, table_name):
    cursor.execute("SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = ?", (table_name,))
    return [row[0].lower() for row in cursor.fetchall()]

def get_all_db_tables(cursor):
    cursor.execute("SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_TYPE = 'BASE TABLE'")
    return [row[0].lower().strip() for row in cursor.fetchall()]

def parse_excel_date(val):
    if not val:
        return None
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime('%Y-%m-%d %H:%M:%S')
    val_str = str(val).strip()
    if val_str.upper() in ('', 'NONE', 'NULL', '-'):
        return None
    for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d.%m.%Y %H:%M:%S'):
        try:
            dt = datetime.datetime.strptime(val_str, fmt)
            return dt.strftime('%Y-%m-%d %H:%M:%S')
        except ValueError:
            continue
    return val_str

def load_excel_to_sql():
    from core.database_sql import init_db
    print("[*] Veritabanı (SQL) kolonlarına varana kadar temizleniyor ve yeni şemayla (İngilizce) oluşturuluyor...")
    init_db()

    if not os.path.exists(DATABASE_DIR):
        print(f"[!] database/ klasoru bulunamadi: {DATABASE_DIR}")
        return

    # Sadece gecersiz/yedek olmayan Excel dosyalarini hedefle
    excel_files = [f for f in os.listdir(DATABASE_DIR) if f.endswith('.xlsx') and not f.startswith('~')]
    if not excel_files:
        print("[!] database/ klasorunde yuklenecek Excel (.xlsx) dosyasi bulunamadi.")
        return

    conn = get_connection()
    cursor = conn.cursor()
    
    try:
        db_tables = get_all_db_tables(cursor)
    except Exception as e:
        db_tables = []
    
    total_inserted = 0
    total_errors = 0
    skipped_records = []

    sheet_map = {
        'pcs': 'pcs', 
        'bilgisayar': 'pcs',
        'queing_machines': 'queing_machines', 
        'siramatik': 'queing_machines', 
        'sıramatik': 'queing_machines', 
        'kiosk': 'queing_machines',
        'tablets': 'tablets', 
        'tabletler': 'tablets', 
        'tablet': 'tablets',
        'printers': 'printers', 
        'yazıcılar': 'printers', 
        'yazıcı': 'printers', 
        'yazici': 'printers',
        'barcode_printers': 'barcode_printers', 
        'barkod yazıcı': 'barcode_printers',
        'barcode_readers': 'barcode_readers', 
        'barkod okuyucu': 'barcode_readers',
        'scanners': 'scanners', 
        'tarayıcı': 'scanners',
        'depot_items': 'depot_items', 
        'depo envanteri': 'depot_items',
        'consumable_items': 'consumable_items', 
        'sarf malzemeleri': 'consumable_items',
        'users': 'users', 
        'kullanıcılar': 'users', 
        'kullanicilar': 'users',
        'shared_areas': 'shared_areas',
        'ortak_alanlar': 'shared_areas',
        'ortak alanlar': 'shared_areas',
        'technical_notes': 'technical_notes',
        'technical notes': 'technical_notes',
        'kodlar': 'technical_notes',
        'closure_notes': 'closure_notes',
        'closure notes': 'closure_notes',
        'kapanis': 'closure_notes',
        'kapanış': 'closure_notes',
        'troubleshooting_notes': 'troubleshooting_notes',
        'troubleshooting notes': 'troubleshooting_notes',
        'sorun-giderme': 'troubleshooting_notes',
        'sorun giderme': 'troubleshooting_notes',
        'mahal_list': 'mahal_list', 
        'mahal listesi': 'mahal_list',
        'list': 'mahal_list',
        'printer_service': 'printer_service',
        'printer service': 'printer_service',
        'printer_service_history': 'printer_service_history',
        'printer service history': 'printer_service_history',
        'monitors': 'monitors',
        'monitörler': 'monitors',
        'ekranlar': 'monitors',
        # --- SISTEM TABLOLARI ---
        'audit_logs': 'audit_logs',
        'audit logs': 'audit_logs',
        'refresh_tokens': 'refresh_tokens',
        'refresh tokens': 'refresh_tokens',
        'user_activity_log': 'user_activity_log',
        'user activity log': 'user_activity_log'
    }

    for excel_file in excel_files:
        filepath = os.path.join(DATABASE_DIR, excel_file)
        print(f"\n==================================================")
        print(f"  ISLENEN DOSYA: {excel_file}")
        print(f"==================================================")
        
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        except Exception as e:
            print(f"  [HATA] Dosya acilamadi: {e}")
            total_errors += 1
            continue

        for sheet_name in wb.sheetnames:
            clean_name = sheet_name.strip().lower()
            table_name = sheet_map.get(clean_name)
            
            # Dinamik Eşleme (Dynamic Mapping Fallback)
            if not table_name:
                if clean_name in db_tables:
                    table_name = clean_name
                elif clean_name.replace(' ', '_') in db_tables:
                    table_name = clean_name.replace(' ', '_')
                elif clean_name.replace('_', ' ') in db_tables:
                    table_name = clean_name.replace('_', ' ')
            
            if not table_name:
                print(f"[-] Eşleme Dışı Sayfa: '{sheet_name}' (Bu sayfa SQL eşleme haritasında tanımlı değil, atlanıyor.)")
                continue

            # Tablo kolonlarini al
            try:
                sql_columns = get_table_columns(cursor, table_name)
            except Exception as e:
                sql_columns = []

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                print(f"  [BOŞ SAYFA] '{sheet_name}' sayfasında veri bulunamadı.")
                continue

            if not sql_columns:
                print(f"  [UYARI] '{sheet_name}' -> SQL'de '{table_name}' tablosu bulunamadi. Dinamik olarak oluşturuluyor...")
                
                # Sütunları belirle
                headers = []
                for h in rows[0]:
                    col = str(h).strip().lower() if h else ''
                    if col == 'saha_stock': col = 'field_stock'
                    elif col == 'arizali_stock': col = 'faulty_stock'
                    elif col == 'kayip_stock': col = 'lost_stock'
                    elif col in ('sahada', 'kurulu'): col = 'on_field'
                    elif col == 'depo': col = 'warehouse'
                    elif col in ('arizali', 'arızalı'): col = 'is_faulty'
                    elif col in ('kayip', 'kayıp', 'mahalsiz'): col = 'without_location'
                    elif col == 'deleted': col = 'is_deleted'
                    headers.append(col)
                valid_headers = [h for h in headers if h and h not in ('id', 'created_at', 'last_edit_date', 'last_edit_user')]
                
                if not valid_headers:
                    print(f"    [HATA] '{sheet_name}' sayfasında geçerli başlık bulunamadı.")
                    continue
                    
                # CREATE TABLE sorgusunu hazırla
                cols_def = []
                for h in valid_headers:
                    if h == 'is_deleted':
                        continue  # We explicitly define it as BIT DEFAULT 0 below
                    cols_def.append(f"[{h}] NVARCHAR(255)")
                
                if table_name == 'users':
                    required_users_cols = ['password_hash', 'display_name', 'role', 'permissions', 'keyos_user', 'keyos_pass', 'bim_user', 'bim_pass', 'session_timeout', 'last_login']
                    for rc in required_users_cols:
                        if rc not in valid_headers:
                            if rc == 'session_timeout':
                                cols_def.append(f"[{rc}] INT DEFAULT 30")
                            elif rc == 'last_login':
                                cols_def.append(f"[{rc}] DATETIME")
                            else:
                                cols_def.append(f"[{rc}] NVARCHAR(255)")

                create_sql = f"CREATE TABLE [{table_name}] (id INT IDENTITY(1,1) PRIMARY KEY, created_at DATETIME DEFAULT GETDATE(), last_edit_date DATETIME, last_edit_user NVARCHAR(255), is_deleted BIT DEFAULT 0, {', '.join(cols_def)})"
                
                try:
                    cursor.execute(create_sql)
                    conn.commit()
                    sql_columns = get_table_columns(cursor, table_name)
                    print(f"    [+] TABLO OLUŞTURULDU: '{table_name}' ({len(valid_headers)} kolon eklendi)")
                except Exception as e:
                    print(f"    [HATA] Tablo oluşturulurken hata: {e}")
                    continue

            # Basliklar
            headers = []
            for h in rows[0]:
                col = str(h).strip().lower() if h else ''
                if col == 'saha_stock': col = 'field_stock'
                elif col == 'arizali_stock': col = 'faulty_stock'
                elif col == 'kayip_stock': col = 'lost_stock'
                elif col in ('sahada', 'kurulu'): col = 'on_field'
                elif col == 'depo': col = 'warehouse'
                elif col in ('arizali', 'arızalı'): col = 'is_faulty'
                elif col in ('kayip', 'kayıp', 'mahalsiz'): col = 'without_location'
                elif col == 'deleted': col = 'is_deleted'
                headers.append(col)
            
            skip_cols = {'id', 'created_at', 'last_edit_date', 'last_edit_user'}
            matched_cols = []
            matched_indices = []
            
            # Dinamik Kolon Ekleme (Excel'de var, SQL'de yoksa otomatik ekle)
            for idx, h in enumerate(headers):
                if h and h not in skip_cols:
                    if h not in sql_columns:
                        try:
                            cursor.execute(f"ALTER TABLE [{table_name}] ADD [{h}] NVARCHAR(255)")
                            conn.commit()
                            sql_columns.append(h)
                            print(f"    [+] YENİ SÜTUN: '{table_name}' tablosuna Excel'den '{h}' eklendi.")
                        except Exception as e:
                            print(f"    [HATA] '{h}' sütunu eklenirken hata: {e}")
                            
                    if h in sql_columns:
                        matched_cols.append(h)
                        matched_indices.append(idx)

            print(f"\n[+] Sayfa Bulundu: '{sheet_name}' -> SQL: '{table_name}'")
            print(f"    - Eşleşen Sütunlar ({len(matched_cols)} adet): {', '.join(matched_cols)}")

            if not matched_cols:
                continue

            insert_count = 0
            error_count = 0

            has_id = 'id' in matched_cols
            
            placeholders = ','.join(['?' for _ in matched_cols])
            col_names = ','.join([f'[{c}]' for c in matched_cols])
            insert_sql = f"INSERT INTO [{table_name}] ({col_names}) VALUES ({placeholders})"

            for row_idx, row in enumerate(rows[1:], start=2):
                try:
                    pc_no_val = None
                    keyos_loc_val = None
                    if table_name == 'pcs':
                        if 'pc_no' in headers:
                            p_idx = headers.index('pc_no')
                            if p_idx < len(row):
                                pc_no_val = str(row[p_idx]).strip() if row[p_idx] is not None else None
                        if 'keyos_location' in headers:
                            k_idx = headers.index('keyos_location')
                            if k_idx < len(row):
                                keyos_loc_val = str(row[k_idx]).strip() if row[k_idx] is not None else None

                    values = []
                    for col_idx in matched_indices:
                        header = matched_cols[matched_indices.index(col_idx)]
                        val = row[col_idx] if col_idx < len(row) else None
                        
                        if table_name == 'pcs' and header == 'ip':
                            val_str = str(val).strip().upper() if val is not None else ""
                            if val_str in ('#N/A', 'NONE', 'NULL', '-') or val_str.startswith('='):
                                val = None

                        if table_name == 'pcs' and header == 'hostname':
                            val_str = str(val).strip().upper() if val is not None else ""
                            if not val or val_str in ('', 'NONE', 'NULL', '-'):
                                if keyos_loc_val and keyos_loc_val not in ('', 'NONE', 'NULL', '-'):
                                    val = keyos_loc_val

                        val_str = str(val).strip().upper() if val is not None else ""
                        
                        if val is None or val_str in ('', 'NONE', 'NULL', '-'):
                            val = None
                        elif header in ('acquisition_date', 'sent_date', 'return_date', 'created_at', 'last_edit_date', 'last_login'):
                            val = parse_excel_date(val)
                        else:
                            if table_name == 'users' and header == 'password_hash':
                                if val and not str(val).startswith(('pbkdf2:', 'scrypt:', 'bcrypt:')):
                                    val = generate_password_hash(str(val))

                            is_bit_col = header in ('on_field', 'warehouse', 'is_faulty', 'without_location', 'pending_installation', 'windows', 'keyos', 'rdp', 'is_deleted', 'in_service', 'hostname_mismatch', 'requires_user', 'has_substitute')
                            if is_bit_col:
                                if val_str in ('ON FIELD', 'SAHADA', 'EVET', '1', 'TRUE', 'VAR', 'AKTIF', 'VERILDI', 'DOGRU', 'DOĞRU'):
                                    val = 1 if (header == 'on_field' or not val_str in ('ON FIELD', 'SAHADA')) else 0
                                    if val_str in ('ON FIELD', 'SAHADA'): val = 1 if header == 'on_field' else 0
                                    else: val = 1
                                elif val_str in ('WAREHOUSE', 'DEPO', 'ARIZALI', 'IS FAULTY', 'BOZUK', 'KAYIP', 'MAHALSIZ', 'WITHOUT LOCATION', 'HAYIR', '0', 'FALSE', 'YOK', 'PASIF', 'VERILMEDI', 'YANLIS', 'YANLIŞ'):
                                    if val_str in ('WAREHOUSE', 'DEPO'): val = 1 if header == 'warehouse' else 0
                                    elif val_str in ('ARIZALI', 'IS FAULTY', 'BOZUK'): val = 1 if header == 'is_faulty' else 0
                                    elif val_str in ('KAYIP', 'MAHALSIZ', 'WITHOUT LOCATION'): val = 1 if header == 'without_location' else 0
                                    else: val = 0
                                else:
                                    try: val = 1 if float(val) > 0 else 0
                                    except: val = 0
                        
                        values.append(val)

                    if all(v is None for v in values):
                        continue

                    if table_name in ('printer_service', 'printer_service_history'):
                        has_real_data = False
                        for col_name in ('acquisition_date', 'sent_date', 'return_date', 'fault_description'):
                            if col_name in matched_cols:
                                val_check = values[matched_cols.index(col_name)]
                                if val_check is not None and str(val_check).strip() not in ('', 'None', 'NULL', '-'):
                                    has_real_data = True
                                    break
                        if not has_real_data:
                            continue

                    unique_col = None
                    if table_name == 'users': unique_col = 'username'
                    elif table_name == 'pcs': unique_col = 'pc_serial'
                    elif table_name == 'queing_machines': unique_col = 'ip'
                    elif table_name == 'tablets': unique_col = 'ip'
                    elif table_name == 'mahal_list': unique_col = 'location_code'
                    elif table_name == 'shared_areas': unique_col = 'name'
                    elif table_name in ('barcode_printers', 'barcode_readers', 'scanners'): unique_col = 'serial_no'
                    elif table_name in ('technical_notes', 'closure_notes', 'troubleshooting_notes'): unique_col = 'title'
                    elif table_name == 'depot_items': unique_col = 'name'

                    if table_name == 'printers':
                        s_idx = matched_cols.index('serial_no') if 'serial_no' in matched_cols else -1
                        m_idx = matched_cols.index('mac') if 'mac' in matched_cols else -1
                        s_val = str(values[s_idx]).strip() if s_idx != -1 and values[s_idx] is not None else ''
                        m_val = str(values[m_idx]).strip() if m_idx != -1 and values[m_idx] is not None else ''
                        
                        if s_val in ('', 'NONE', 'NULL', '-') and m_val in ('', 'NONE', 'NULL', '-'):
                            pass
                        else:
                            query = "SELECT COUNT(*) FROM [printers] WHERE (is_deleted = 0 OR is_deleted IS NULL)"
                            params = []
                            if s_val not in ('', 'NONE', 'NULL', '-'):
                                query += " AND TRIM(UPPER([serial_no])) = TRIM(UPPER(?))"
                                params.append(s_val)
                            if m_val not in ('', 'NONE', 'NULL', '-'):
                                query += " AND TRIM(UPPER([mac])) = TRIM(UPPER(?))"
                                params.append(m_val)
                            if params:
                                cursor.execute(query, params)
                                if cursor.fetchone()[0] > 0:
                                    skipped_records.append(f"Yazıcılar: Seri No ({s_val}) veya Mac ({m_val}) zaten var, eklenmedi.")
                                    continue

                    if unique_col and unique_col in matched_cols:
                        u_idx = matched_cols.index(unique_col)
                        u_val = values[u_idx]
                        if u_val and str(u_val).strip() not in ('', 'NONE', 'NULL', '-'):
                            is_deleted_clause = " AND (is_deleted = 0 OR is_deleted IS NULL)" if table_name != 'users' else ""
                            cursor.execute(f"SELECT COUNT(*) FROM [{table_name}] WHERE TRIM(UPPER([{unique_col}])) = TRIM(UPPER(?)){is_deleted_clause}", (u_val,))
                            if cursor.fetchone()[0] > 0:
                                skipped_records.append(f"{table_name}: '{u_val}' ({unique_col}) zaten var, eklenmedi.")
                                continue

                    # INSERT CALISTIR (IDENTITY_INSERT DESTEGI ILE)
                    if has_id:
                        try:
                            cursor.execute(f"SET IDENTITY_INSERT [{table_name}] ON")
                            cursor.execute(insert_sql, values)
                            cursor.execute(f"SET IDENTITY_INSERT [{table_name}] OFF")
                        except Exception as insert_e:
                            try:
                                cursor.execute(f"SET IDENTITY_INSERT [{table_name}] OFF")
                            except Exception as set_identity_off_ex:
                                print(f"[Excel Load Identity Off Error] {set_identity_off_ex}")
                            
                            # ID eklerken hata olduysa (belki ID identity degildir), ID'siz normal ekleme yapalim:
                            insert_dict = dict(zip(matched_cols, values))
                            if 'id' in insert_dict:
                                del insert_dict['id']
                            fallback_cols = list(insert_dict.keys())
                            fallback_vals = list(insert_dict.values())
                            fb_placeholders = ','.join(['?' for _ in fallback_cols])
                            fb_col_names = ','.join([f'[{c}]' for c in fallback_cols])
                            fb_sql = f"INSERT INTO [{table_name}] ({fb_col_names}) VALUES ({fb_placeholders})"
                            cursor.execute(fb_sql, fallback_vals)
                    else:
                        cursor.execute(insert_sql, values)
                        
                    insert_count += 1
                except Exception as e:
                    error_count += 1
                    if error_count <= 5:
                        print(f"    [HATA] Satir {row_idx}: {e}")

            conn.commit()
            total_inserted += insert_count
            total_errors += error_count
            print(f"    [✓] {insert_count} kayit islendi. ({error_count} hata)")

        wb.close()

    # --- KRITIK LOGIN KORUMASI: Varsayilan Yonetici Seeding ---
    try:
        admin_default_pass = os.getenv("ADMIN_DEFAULT_PASS", "change_me_immediately")
        admin_pw_hash = generate_password_hash(admin_default_pass)
        cursor.execute(
            "IF EXISTS (SELECT * FROM users WHERE username='vefa') "
            "UPDATE users SET role='ADMIN' WHERE username='vefa' "
            "ELSE INSERT INTO users (username, password_hash, display_name, role) "
            "VALUES ('vefa', ?, 'Vefa', 'ADMIN')",
            (admin_pw_hash,)
        )
        conn.commit()
        print("\n  [✓] Varsayilan Yonetici hesabi (vefa) guncellendi/olusturuldu.")
    except Exception as e:
        print(f"\n  [!] Yonetici seed hatasi: {e}")

    conn.close()
    print(f"\n==================================================")
    print(f"  TOPLAM: {total_inserted} kayit yuklendi, {total_errors} hata.")
    print(f"==================================================")

    if skipped_records:
        print("\n=== ATLANAN / EKLENMEYEN KAYITLAR RAPORU ===")
        # En fazla ilk 500 atlanan kaydı yazdır (Terminal şişmemesi için)
        for msg in skipped_records[:500]:
            print(f"- {msg}")
        if len(skipped_records) > 500:
            print(f"... ve {len(skipped_records) - 500} kayit daha atlandi.")
        print("==================================================\n")

if __name__ == "__main__":
    load_excel_to_sql()
