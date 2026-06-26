import os
import sys
# Path patch for sub-folder execution
project_root = os.path.dirname(os.path.abspath(__file__))
if os.path.basename(project_root) == 'tools':
    project_root = os.path.dirname(project_root)
if project_root not in sys.path:
    sys.path.append(project_root)
os.chdir(project_root)

# -*- coding: utf-8 -*-
"""
IT_INVENTORY v13.0 — Excel'den SQL Server'a Toplu Veri Güncelleme (UPSERT) Motoru
"""
import os
import sys
import pyodbc
import openpyxl
import datetime
from dotenv import load_dotenv
from werkzeug.security import generate_password_hash

BASE_DIR = project_root
env_path = os.path.join(BASE_DIR, ".env")
if not os.path.exists(env_path):
    env_path = os.path.join(BASE_DIR, "tools", ".env")
load_dotenv(env_path, override=True)

DB_SERVER = os.getenv("DB_SERVER", ".\\SQLEXPRESS").strip()
DB_NAME = os.getenv("DB_NAME", "IT_INVENTORY").strip()
DB_USER = os.getenv("DB_USER", "").strip()
DB_PASS = os.getenv("DB_PASS", "").strip()

DATABASE_DIR = os.path.join(BASE_DIR, "database")

def get_connection():
    if DB_USER and DB_PASS:
        conn_str = f"DRIVER={{SQL Server}};SERVER={DB_SERVER};DATABASE={DB_NAME};UID={DB_USER};PWD={DB_PASS};"
    else:
        conn_str = f"DRIVER={{SQL Server}};SERVER={DB_SERVER};DATABASE={DB_NAME};Trusted_Connection=yes;"
    return pyodbc.connect(conn_str, timeout=10)

def get_table_columns(cursor, table_name):
    cursor.execute("SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = ?", (table_name,))
    return [row[0].lower() for row in cursor.fetchall()]

def parse_excel_date(val):
    if not val:
        return None
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime('%Y-%m-%d %H:%M:%S')
    val_str = str(val).strip()
    if val_str.upper() in ('', 'NONE', 'NULL', '-'):
        return None
    for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d.%m.%Y %H:%M:%S'):
        try:
            dt = datetime.datetime.strptime(val_str, fmt)
            return dt.strftime('%Y-%m-%d %H:%M:%S')
        except ValueError:
            continue
    return val_str

def update_excel_to_sql():
    from core.database_sql import init_db
    print("[*] Veritabanı (SQL) şema gereksinimleri kontrol ediliyor...")
    init_db()

    if not os.path.exists(DATABASE_DIR):
        print(f"[!] database/ klasoru bulunamadi: {DATABASE_DIR}")
        return

    excel_files = [f for f in os.listdir(DATABASE_DIR) if f.endswith('.xlsx') and not f.startswith('~')]
    if not excel_files:
        print("[!] database/ klasorunde guncellenecek Excel (.xlsx) dosyasi bulunamadi.")
        return

    conn = get_connection()
    cursor = conn.cursor()
    
    total_updated = 0
    total_inserted = 0
    total_errors = 0
    skipped_records = []

    sheet_map = {
        'pcs': 'pcs', 'bilgisayar': 'pcs',
        'queing_machines': 'queing_machines', 'siramatik': 'queing_machines', 'sıramatik': 'queing_machines', 'kiosk': 'queing_machines',
        'tablets': 'tablets', 'tabletler': 'tablets', 'tablet': 'tablets',
        'printers': 'printers', 'yazıcılar': 'printers', 'yazıcı': 'printers', 'yazici': 'printers',
        'barcode_printers': 'barcode_printers', 'barkod yazıcı': 'barcode_printers',
        'barcode_readers': 'barcode_readers', 'barkod okuyucu': 'barcode_readers',
        'scanners': 'scanners', 'tarayıcı': 'scanners',
        'depot_items': 'depot_items', 'depo envanteri': 'depot_items',
        'consumable_items': 'consumable_items', 'sarf malzemeleri': 'consumable_items',
        'users': 'users', 'kullanıcılar': 'users', 'kullanicilar': 'users',
        'shared_areas': 'shared_areas', 'ortak_alanlar': 'shared_areas', 'ortak alanlar': 'shared_areas',
        'technical_notes': 'technical_notes', 'technical notes': 'technical_notes', 'kodlar': 'technical_notes',
        'closure_notes': 'closure_notes', 'closure notes': 'closure_notes', 'kapanis': 'closure_notes', 'kapanış': 'closure_notes',
        'troubleshooting_notes': 'troubleshooting_notes', 'troubleshooting notes': 'troubleshooting_notes', 'sorun-giderme': 'troubleshooting_notes', 'sorun giderme': 'troubleshooting_notes',
        'mahal_list': 'mahal_list', 'mahal listesi': 'mahal_list', 'list': 'mahal_list',
        'printer_service': 'printer_service', 'printer service': 'printer_service',
        'printer_service_history': 'printer_service_history', 'printer service history': 'printer_service_history',
        'monitors': 'monitors', 'monitörler': 'monitors', 'ekranlar': 'monitors'
    }

    for excel_file in excel_files:
        filepath = os.path.join(DATABASE_DIR, excel_file)
        print(f"\n==================================================")
        print(f"  ISLENEN DOSYA (GÜNCELLEME): {excel_file}")
        print(f"==================================================")
        
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        except Exception as e:
            print(f"  [HATA] Dosya acilamadi: {e}")
            total_errors += 1
            continue

        for sheet_name in wb.sheetnames:
            clean_name = sheet_name.strip().lower()
            table_name = sheet_map.get(clean_name)
            
            # Dinamik eşleşme (Eğer eşleme haritasında yoksa, adını doğrudan tablo ismi olarak dene)
            if not table_name:
                cursor.execute("SELECT COUNT(*) FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = ?", (clean_name,))
                if cursor.fetchone()[0] > 0:
                    table_name = clean_name
                else:
                    # Export edilirken 31 karaktere kirpilmis olabilir, oyleyse eslestir.
                    cursor.execute("SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES")
                    all_tables = [row[0] for row in cursor.fetchall()]
                    for t in all_tables:
                        if t[:31].lower() == clean_name:
                            table_name = t
                            break
                    if not table_name:
                        print(f"[-] Eşleme Dışı Sayfa: '{sheet_name}' (Veritabanında böyle bir tablo bulunamadı, atlanıyor.)")
                        continue

            try:
                sql_columns = get_table_columns(cursor, table_name)
            except Exception as e:
                sql_columns = []

            if not sql_columns:
                print(f"  [HATA] SQL'de '{table_name}' tablosu bulunamadi. Atlanıyor.")
                continue

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                print(f"  [BOŞ SAYFA] '{sheet_name}' sayfasında veri bulunamadı.")
                continue

            headers = []
            for h in rows[0]:
                col = str(h).strip().lower() if h else ''
                if col == 'saha_stock': col = 'field_stock'
                elif col == 'arizali_stock': col = 'faulty_stock'
                elif col == 'kayip_stock': col = 'lost_stock'
                elif col in ('sahada', 'kurulu'): col = 'on_field'
                elif col == 'depo': col = 'warehouse'
                elif col in ('arizali', 'arızalı'): col = 'is_faulty'
                elif col in ('kayip', 'kayıp', 'mahalsiz'): col = 'without_location'
                elif col == 'deleted': col = 'is_deleted'
                headers.append(col)
            
            skip_cols = {'id', 'created_at', 'last_edit_date', 'last_edit_user'}
            matched_cols = []
            matched_indices = []
            
            # Dinamik Kolon Ekleme
            for idx, h in enumerate(headers):
                if h and h not in skip_cols:
                    if h not in sql_columns:
                        try:
                            cursor.execute(f"ALTER TABLE [{table_name}] ADD [{h}] NVARCHAR(255)")
                            conn.commit()
                            sql_columns.append(h)
                            print(f"    [+] YENİ SÜTUN: '{table_name}' tablosuna Excel'den '{h}' eklendi.")
                        except Exception as e:
                            print(f"    [HATA] '{h}' sütunu eklenirken hata: {e}")
                    if h in sql_columns:
                        matched_cols.append(h)
                        matched_indices.append(idx)

            print(f"\n[+] Sayfa Bulundu: '{sheet_name}' -> SQL: '{table_name}'")

            if not matched_cols:
                continue

            update_count = 0
            insert_count = 0
            error_count = 0

            # Benzersiz Kolon Belirleme
            unique_col = None
            if table_name == 'users': unique_col = 'username'
            elif table_name == 'pcs': unique_col = 'pc_serial'
            elif table_name == 'queing_machines': unique_col = 'ip'
            elif table_name == 'tablets': unique_col = 'ip'
            elif table_name == 'mahal_list': unique_col = 'location_code'
            elif table_name == 'shared_areas': unique_col = 'name'
            elif table_name in ('barcode_printers', 'barcode_readers', 'scanners'): unique_col = 'serial_no'
            elif table_name in ('technical_notes', 'closure_notes', 'troubleshooting_notes'): unique_col = 'title'
            elif table_name == 'depot_items': unique_col = 'name'
            elif table_name == 'printers': unique_col = 'serial_no'
            elif 'id' in matched_cols: unique_col = 'id'
            
            if not unique_col:
                # Eger hicbiri yoksa ve id diye bir kolon da yoksa bile id diye zorlayalim
                try:
                    cursor.execute(f"SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME='{table_name}' AND COLUMN_NAME='id'")
                    if cursor.fetchone(): unique_col = 'id'
                except Exception as ex_id:
                    print(f"[Excel Update Unique ID Search Exception] {ex_id}")

            if not unique_col or unique_col not in matched_cols:
                print(f"  [UYARI] '{table_name}' için benzersiz bir sütun ({unique_col}) bulunamadı. Sadece Ekleme yapılacak.")

            for row_idx, row in enumerate(rows[1:], start=2):
                try:
                    values_dict = {}
                    for col_idx in matched_indices:
                        header = matched_cols[matched_indices.index(col_idx)]
                        val = row[col_idx] if col_idx < len(row) else None
                        
                        if table_name == 'pcs' and header == 'ip':
                            val_str = str(val).strip().upper() if val is not None else ""
                            if val_str in ('#N/A', 'NONE', 'NULL', '-') or val_str.startswith('='):
                                val = None

                        val_str = str(val).strip().upper() if val is not None else ""
                        
                        if val is None or val_str in ('', 'NONE', 'NULL', '-'):
                            val = None
                        elif header in ('acquisition_date', 'sent_date', 'return_date'):
                            val = parse_excel_date(val)
                        else:
                            if table_name == 'users' and header == 'password_hash':
                                if val and not str(val).startswith(('pbkdf2:', 'scrypt:', 'bcrypt:')):
                                    val = generate_password_hash(str(val))

                            is_bit_col = header in ('on_field', 'warehouse', 'is_faulty', 'without_location', 'pending_installation', 'windows', 'keyos', 'rdp', 'is_deleted', 'in_service', 'hostname_mismatch', 'requires_user', 'has_substitute')
                            if is_bit_col:
                                if val_str in ('ON FIELD', 'SAHADA', 'EVET', '1', 'TRUE', 'VAR', 'AKTIF', 'VERILDI', 'DOGRU', 'DOĞRU'):
                                    val = 1 if (header == 'on_field' or not val_str in ('ON FIELD', 'SAHADA')) else 0
                                    if val_str in ('ON FIELD', 'SAHADA'): val = 1 if header == 'on_field' else 0
                                    else: val = 1
                                elif val_str in ('WAREHOUSE', 'DEPO', 'ARIZALI', 'IS FAULTY', 'BOZUK', 'KAYIP', 'MAHALSIZ', 'WITHOUT LOCATION', 'HAYIR', '0', 'FALSE', 'YOK', 'PASIF', 'VERILMEDI', 'YANLIS', 'YANLIŞ'):
                                    if val_str in ('WAREHOUSE', 'DEPO'): val = 1 if header == 'warehouse' else 0
                                    elif val_str in ('ARIZALI', 'IS FAULTY', 'BOZUK'): val = 1 if header == 'is_faulty' else 0
                                    elif val_str in ('KAYIP', 'MAHALSIZ', 'WITHOUT LOCATION'): val = 1 if header == 'without_location' else 0
                                    else: val = 0
                                else:
                                    try: val = 1 if float(val) > 0 else 0
                                    except: val = 0
                        
                        values_dict[header] = val

                    if all(v is None for v in values_dict.values()):
                        continue

                    u_val = values_dict.get(unique_col) if unique_col else None
                    record_exists = False
                    existing_row = None
                    used_match_col = None

                    # 1. ÖNCELİK: Eğer satırda 'id' varsa, KESİNLİKLE ID'den kontrol et (Çünkü Excel'de ID varsa, DB'den gelmiştir)
                    if 'id' in matched_cols:
                        id_val = values_dict.get('id')
                        if id_val and str(id_val).strip() not in ('', 'NONE', 'NULL', '-'):
                            is_deleted_clause = " AND (is_deleted = 0 OR is_deleted IS NULL)" if table_name != 'users' else ""
                            try:
                                cursor.execute(f"SELECT * FROM [{table_name}] WHERE id = ?{is_deleted_clause}", (id_val,))
                                existing_row = cursor.fetchone()
                                if existing_row:
                                    record_exists = True
                                    used_match_col = f"id={id_val}"
                            except Exception as id_e:
                                print(f"[Excel Update ID Match Error] {id_e}")

                    # 2. ÖNCELİK: ID'den bulunamadıysa veya ID yoksa, unique_col (serial_no, username vb.) ile kontrol et
                    if not record_exists and unique_col and unique_col != 'id' and u_val and str(u_val).strip() not in ('', 'NONE', 'NULL', '-'):
                        is_deleted_clause = " AND (is_deleted = 0 OR is_deleted IS NULL)" if table_name != 'users' else ""
                        cursor.execute(f"SELECT * FROM [{table_name}] WHERE TRIM(UPPER([{unique_col}])) = TRIM(UPPER(?)){is_deleted_clause}", (u_val,))
                        existing_row = cursor.fetchone()
                        if existing_row:
                            record_exists = True
                            used_match_col = f"{unique_col}={u_val}"
                        
                    if record_exists and existing_row:
                            # Degisiklikleri karsilastir ve sadece degisenleri UPDATE et
                            db_columns = [column[0].lower() for column in cursor.description]
                            db_dict = dict(zip(db_columns, existing_row))
                            
                            updates_needed = {}
                            for h, excel_val in values_dict.items():
                                db_val = db_dict.get(h)
                                
                                # Tip dönusumleri yaparak temiz karşilaştirma (ör. int -> string, vs.)
                                e_val = str(excel_val).strip() if excel_val is not None else ""
                                d_val = str(db_val).strip() if db_val is not None else ""
                                
                                # Python datetime formatları SQL server ile milisaniye bazında uyuşmayabilir, basitçe ilk 19 karakter
                                if h in ('acquisition_date', 'sent_date', 'return_date'):
                                    e_val = e_val[:19]
                                    d_val = d_val[:19]
                                
                                # Eğer Excel'de boşsa ve veritabanında da boş/Yok ise, veya değerler eşitse atla.
                                if e_val == d_val:
                                    continue
                                
                                # Float/Int ".0" sorunları
                                if e_val.endswith('.0') and e_val[:-2] == d_val: continue
                                if d_val.endswith('.0') and d_val[:-2] == e_val: continue
                                
                                updates_needed[h] = excel_val

                            if updates_needed:
                                set_clauses = ', '.join([f"[{k}] = ?" for k in updates_needed.keys()])
                                update_values = list(updates_needed.values())
                                update_values.append(u_val) # WHERE için
                                
                                update_sql = f"UPDATE [{table_name}] SET {set_clauses} WHERE TRIM(UPPER([{unique_col}])) = TRIM(UPPER(?)){is_deleted_clause}"
                                cursor.execute(update_sql, update_values)
                                update_count += 1
                                print(f"      [Güncellendi] {table_name} -> {unique_col}={u_val} | Değişenler: {list(updates_needed.keys())}")
                            else:
                                # Değişiklik yok
                                pass

                    if not record_exists:
                        # Kayıt bulunamadı, INSERT yap
                        insert_dict = {k: v for k, v in values_dict.items() if v is not None and str(v).strip() != ''}
                        
                        try_identity_insert = False
                        if 'id' in insert_dict:
                            try_identity_insert = True
                            
                        insert_cols = list(insert_dict.keys())
                        placeholders = ','.join(['?' for _ in insert_cols])
                        col_names = ','.join([f'[{c}]' for c in insert_cols])
                        insert_values = list(insert_dict.values())
                        
                        try:
                            if try_identity_insert:
                                cursor.execute(f"SET IDENTITY_INSERT [{table_name}] ON")
                                insert_sql = f"INSERT INTO [{table_name}] ({col_names}) VALUES ({placeholders})"
                                cursor.execute(insert_sql, insert_values)
                                cursor.execute(f"SET IDENTITY_INSERT [{table_name}] OFF")
                            else:
                                insert_sql = f"INSERT INTO [{table_name}] ({col_names}) VALUES ({placeholders})"
                                cursor.execute(insert_sql, insert_values)
                        except Exception as insert_ex:
                            # Eger identity hatasi verirse (veya ID identity degilse) fallback
                            try:
                                cursor.execute(f"SET IDENTITY_INSERT [{table_name}] OFF")
                            except Exception as set_identity_off_ex:
                                print(f"[Excel Update Identity Off Error] {set_identity_off_ex}")
                            
                            if 'id' in insert_dict:
                                del insert_dict['id']
                                insert_cols = list(insert_dict.keys())
                                placeholders = ','.join(['?' for _ in insert_cols])
                                col_names = ','.join([f'[{c}]' for c in insert_cols])
                                insert_values = list(insert_dict.values())
                                insert_sql = f"INSERT INTO [{table_name}] ({col_names}) VALUES ({placeholders})"
                                cursor.execute(insert_sql, insert_values)
                            else:
                                raise insert_ex
                        
                        insert_count += 1
                        print(f"      [Eklendi] {table_name} -> Yeni Kayıt eklendi (Unique Val: {u_val})")

                except Exception as e:
                    error_count += 1
                    if error_count <= 5:
                        print(f"    [HATA] Satir {row_idx}: {e}")

            conn.commit()
            total_updated += update_count
            total_inserted += insert_count
            total_errors += error_count
            print(f"    [✓] {update_count} kayıt güncellendi, {insert_count} yeni kayıt eklendi. ({error_count} hata)")

        wb.close()

    conn.close()
    print(f"\n==================================================")
    print(f"  TOPLAM: {total_updated} kayıt güncellendi, {total_inserted} yeni kayıt eklendi, {total_errors} hata.")
    print(f"==================================================")

if __name__ == "__main__":
    update_excel_to_sql()
